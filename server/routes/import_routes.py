# server/routes/import_routes.py
# Import blueprint: file upload, history, template download, and data export
import io
import json
import secrets
import sqlite3
import threading
import time
import zipfile
from datetime import datetime, timedelta, timezone

from flask import Blueprint, request, jsonify, send_file, current_app
from flask_jwt_extended import jwt_required, get_jwt_identity

from services.database import SQLITE_PROVIDER
from services.database_adapter import DatabaseAdapter
from services.import_service import (
    DAILY_IMPORT_HEADERS,
    DREAM_IMPORT_HEADERS,
    IMPORTANT_DAY_IMPORT_HEADERS,
    THOUGHT_RECORD_IMPORT_HEADERS,
    PACKAGE_TYPE,
    PACKAGE_FORMAT_VERSION,
    PORTABILITY_CONTRACT,
    validate_file,
    parse_import_file,
    preview_import_entries,
    commit_import_preview,
    ensure_history_table,
    ensure_import_sessions_table,
    ensure_import_jobs_table,
    ensure_export_history_table,
    create_import_session,
    get_import_session,
    discard_import_session,
    mark_import_session_consumed,
    record_import_history,
    create_pending_import_history,
    finalise_import_history,
    record_export_history,
    get_import_history,
)
from services.media_storage import delete_image, read_media_bytes
from services.sql_compat import in_placeholders

import_bp = Blueprint('import', __name__)

_IMPORT_JOB_PROGRESS: dict[str, dict] = {}
_ACTIVE_IMPORT_JOB_THREADS: set[str] = set()
_IMPORT_JOBS_LOCK = threading.Lock()


def _update_import_job(job_id: str, **changes) -> None:
    with _IMPORT_JOBS_LOCK:
        job = _IMPORT_JOB_PROGRESS.setdefault(job_id, {})
        job.update(changes)
        job['updated_at'] = datetime.now(timezone.utc).strftime('%Y-%m-%dT%H:%M:%SZ')
        job['last_progress_at'] = time.monotonic()


def _get_import_job(job_id: str, user_id: int) -> dict | None:
    conn = get_db()
    ensure_import_jobs_table(conn)
    row = conn.execute(
        '''SELECT id, user_id, import_session_id, status, processed, total, percent,
                  message, error, result_json, created_at, updated_at
           FROM import_jobs WHERE id = ? AND user_id = ?''',
        (job_id, user_id),
    ).fetchone()
    conn.close()
    if not row:
        return None

    public_job = dict(row)
    public_job.pop('user_id', None)
    result_json = public_job.pop('result_json', None)
    if result_json:
        try:
            public_job['result'] = json.loads(result_json)
        except (TypeError, ValueError):
            public_job['result'] = None

    with _IMPORT_JOBS_LOCK:
        live_progress = _IMPORT_JOB_PROGRESS.get(job_id)
        if live_progress:
            public_job.update({
                key: value for key, value in live_progress.items()
                if key != 'last_progress_at'
            })
            public_job['is_delayed'] = (
                public_job['status'] == 'running'
                and time.monotonic() - live_progress['last_progress_at'] > 120
            )
        else:
            public_job['is_delayed'] = False
    return public_job


def _persist_import_job(conn: sqlite3.Connection, job_id: str, **changes) -> None:
    if not changes:
        return
    changes['updated_at'] = datetime.now(timezone.utc).strftime('%Y-%m-%dT%H:%M:%SZ')
    assignments = ', '.join(f'{column} = ?' for column in changes)
    conn.execute(
        f'UPDATE import_jobs SET {assignments} WHERE id = ?',
        (*changes.values(), job_id),
    )
    conn.commit()


def _build_commit_response(
    result: dict,
    *,
    status: str,
    warnings: list[str],
    import_id: int,
) -> dict:
    return {
        'status': status,
        'message': 'Import successful.',
        'summary': {
            'inserted_daily': result['inserted_daily'],
            'skipped_daily': result['skipped_daily'],
            'inserted_dreams': result['inserted_dreams'],
            'skipped_dreams': result['skipped_dreams'],
            'inserted_important_days': result.get('inserted_important_days', 0),
            'inserted_thought_records': result.get('inserted_thought_records', 0),
            'duplicate_dates_daily': result['duplicate_dates_daily'],
            'duplicate_dates_dreams': result['duplicate_dates_dreams'],
        },
        'duplicate_entries': [],
        'warnings': warnings,
        'errors': [],
        'import_id': import_id,
        'import_session_id': None,
    }


def _import_inserted_total(result: dict) -> int:
    return (
        int(result.get('inserted_daily', 0) or 0)
        + int(result.get('inserted_dreams', 0) or 0)
        + int(result.get('inserted_important_days', 0) or 0)
        + int(result.get('inserted_thought_records', 0) or 0)
    )


@import_bp.route('/import/session/<session_id>', methods=['DELETE'])
@jwt_required()
def cancel_import_session(session_id: str):
    user_id = int(get_jwt_identity())
    conn = get_db()
    removed = discard_import_session(conn, user_id=user_id, session_id=session_id)
    conn.close()
    if not removed:
        return jsonify({'error': 'Import review session not found.'}), 404
    return '', 204


def get_db():
    """Get database connection."""
    return _database_adapter().open(
        timeout=30,
        journal_mode_wal=True,
    )


def _database_adapter() -> DatabaseAdapter:
    return current_app.config['DATABASE_ADAPTER']


def _database_provider() -> str:
    return current_app.config.get('DATABASE_PROVIDER', SQLITE_PROVIDER)


def _build_entry_asset_ref(entry_type: str, row: sqlite3.Row) -> str:
    safe_date = str(row['entry_date'] or '').replace('-', '')
    return f"{entry_type}_{safe_date}_{row['entry_number']}_{row['id']}"


def _build_important_day_asset_ref(row: sqlite3.Row) -> str:
    safe_date = str(row['starts_on'] or '').replace('-', '')
    return f"important_day_{safe_date}_{row['id']}"


def _image_filename_for_storage_key(storage_key: str | None) -> str:
    if not storage_key:
        return ''
    return storage_key.rsplit('/', 1)[-1]


def _package_attachment_filename(attachment_row: sqlite3.Row, index: int) -> str:
    base_name = _image_filename_for_storage_key(attachment_row['storage_key'])
    if not base_name:
        original_name = str(attachment_row['original_filename'] or '').strip()
        base_name = original_name.rsplit('/', 1)[-1].rsplit('\\', 1)[-1]
    if not base_name:
        base_name = f'attachment-{index}'
    return f'{index:02d}_{base_name}'


def _load_attachment_export_rows(
    conn: sqlite3.Connection,
    *,
    user_id: int,
    entry_type: str,
    entry_ids: list[int],
) -> dict[int, list[sqlite3.Row]]:
    if not entry_ids:
        return {}

    placeholders = in_placeholders(entry_ids, _database_provider(), start=3)
    rows = conn.execute(
        f'''
        SELECT entry_id, asset_role, storage_key, original_filename, mime_type,
               file_size_bytes, sort_order
        FROM entry_assets
        WHERE user_id = ? AND entry_type = ? AND entry_id IN ({placeholders})
        ORDER BY entry_id ASC, sort_order ASC, id ASC
        ''',
        (user_id, entry_type, *entry_ids),
    ).fetchall()

    grouped: dict[int, list[sqlite3.Row]] = {}
    for row in rows:
        grouped.setdefault(int(row['entry_id']), []).append(row)
    return grouped


# ---------------------------------------------------------------------------
# POST /api/import/upload
# ---------------------------------------------------------------------------

@import_bp.route('/import/upload', methods=['POST'])
@jwt_required()
def upload_import():
    """
    Accept an Excel workbook (.xlsx) or export package (.zip), validate it, parse entries,
    and either import immediately or stage duplicate review before commit.

    Multipart form field: ``file``

    Success response 200:
    {
      "status": "success" | "review_required",
      "summary": {
        ...
      },
      "duplicate_entries": [dict, ...],
      "import_session_id": str | null,
      "warnings": [str, ...],
      "import_id": int | null
    }

    Error response 400 / 422:
    {
      "status": "error",
      "errors": [str, ...]
    }
    """
    user_id = int(get_jwt_identity())

    # --- File presence check ---
    if 'file' not in request.files:
        return jsonify({'status': 'error', 'errors': ['No file part in request.']}), 400

    uploaded_file = request.files['file']
    if not uploaded_file.filename:
        return jsonify({'status': 'error', 'errors': ['No file selected.']}), 400

    # Read the file bytes first so we can validate the size accurately
    file_bytes = uploaded_file.read()
    file_size = len(file_bytes)
    filename = uploaded_file.filename
    content_type = uploaded_file.content_type or ''
    source = str(request.form.get('source') or 'aidiary').strip().lower()

    # --- Structural validation ---
    errors = validate_file(filename, content_type, file_size, source=source)
    if errors:
        return jsonify({'status': 'error', 'errors': errors}), 422

    # --- Parse workbook / package ---
    try:
        parsed = parse_import_file(file_bytes, filename=filename, source=source)
    except ValueError as exc:
        current_app.logger.warning('Import parse error for user %s: %s', user_id, exc)
        message = (
            str(exc)
            if source != 'aidiary'
            else 'The file could not be parsed. Please ensure it is a valid .xlsx workbook or .zip export package.'
        )
        return jsonify({'status': 'error', 'errors': [message]}), 422
    except RuntimeError as exc:
        # pandas / openpyxl not installed
        current_app.logger.error('Import dependency missing: %s', exc)
        return jsonify({'status': 'error', 'errors': ['Workbook import is not available on this server.']}), 500

    parse_errors: list[str] = parsed.get('errors', [])
    if parse_errors:
        return jsonify({'status': 'error', 'errors': parse_errors}), 422

    parse_warnings: list[str] = parsed.get('warnings', [])

    # Reject if nothing parseable was found
    if (
        not parsed.get('daily')
        and not parsed.get('dreams')
        and not parsed.get('important_days')
        and not parsed.get('thought_records')
    ):
        all_warnings = parse_warnings + ['No valid entries found in the uploaded file.']
        conn = get_db()
        ensure_history_table(conn)
        import_id = record_import_history(
            conn, user_id, filename, file_size,
            {'inserted_daily': 0, 'skipped_daily': 0,
             'inserted_dreams': 0, 'skipped_dreams': 0},
            all_warnings, status='empty',
        )
        conn.close()
        return jsonify({
            'status': 'empty',
            'summary': {
                'inserted_daily': 0,
                'skipped_daily': 0,
                'inserted_dreams': 0,
                'skipped_dreams': 0,
                'inserted_important_days': 0,
                'inserted_thought_records': 0,
                'duplicate_dates_daily': [],
                'duplicate_dates_dreams': [],
            },
            'warnings': all_warnings,
            'duplicate_entries': [],
            'import_id': import_id,
        }), 200

    # --- Stage or insert entries ---
    conn = get_db()
    ensure_history_table(conn)
    ensure_import_sessions_table(conn)

    try:
        preview = preview_import_entries(conn, user_id, parsed)
    except Exception as exc:
        conn.close()
        current_app.logger.error('Import insert failed: %s', exc)
        return jsonify({'status': 'error', 'errors': ['Database error during import.']}), 500

    duplicate_rows = preview.get('duplicate_rows', [])
    public_duplicate_rows = [
        {
            'row_id': row['row_id'],
            'entry_type': row['entry_type'],
            'entry_date': row['entry_date'],
            'title': row['title'],
            'reason': row['reason'],
            'content_preview': row['content_preview'],
        }
        for row in duplicate_rows
    ]

    public_review_rows = [
        {
            'row_id': row['_review_row_id'],
            'entry_type': 'daily',
            'entry_date': row['entry_date'],
            'title': row['title'],
            'content_preview': row.get('user_message', '')[:160],
            'mood': row.get('mood', ''),
            'is_duplicate': False,
            'attachment_count': len(row.get('import_attachment_files', [])),
            'source_record_kind': row.get('source_record_kind', 'authored'),
        }
        for row in preview.get('ready_daily_rows', [])
    ] + [
        {
            'row_id': row['_review_row_id'],
            'entry_type': 'dream',
            'entry_date': row['entry_date'],
            'title': row['title'],
            'content_preview': row.get('plot', '')[:160],
            'mood': row.get('emotion', ''),
            'is_duplicate': False,
            'attachment_count': len(row.get('import_attachment_files', [])),
            'source_record_kind': row.get('source_record_kind', 'authored'),
        }
        for row in preview.get('ready_dream_rows', [])
    ] + [
        {
            'row_id': row['_review_row_id'],
            'entry_type': 'important_day',
            'entry_date': row['starts_on'],
            'title': row['label'],
            'content_preview': row.get('note', '')[:160],
            'mood': '',
            'is_duplicate': False,
            'attachment_count': 1 if row.get('import_image_path') else 0,
            'source_record_kind': 'important_day',
        }
        for row in preview.get('ready_important_day_rows', [])
    ] + [
        {
            'row_id': row['_review_row_id'],
            'entry_type': 'thought_record',
            'entry_date': row['record_date'],
            'title': row['title'],
            'content_preview': (row.get('situation') or row.get('balanced_thought') or '')[:160],
            'mood': '',
            'is_duplicate': False,
            'attachment_count': 0,
            'source_record_kind': 'thought_record',
        }
        for row in preview.get('ready_thought_record_rows', [])
    ] + [
        {
            'row_id': row['row_id'],
            'entry_type': row['entry_type'],
            'entry_date': row['entry_date'],
            'title': row['title'],
            'content_preview': row['content_preview'],
            'mood': row.get('row_data', {}).get('mood', ''),
            'is_duplicate': True,
            'attachment_count': len(row.get('row_data', {}).get('import_attachment_files', [])),
            'source_record_kind': row.get('row_data', {}).get('source_record_kind', 'authored'),
        }
        for row in duplicate_rows
    ]

    if duplicate_rows or source != 'aidiary' or parsed.get('important_days') or parsed.get('thought_records'):
        payload = {
            'parse_warnings': parse_warnings,
            **preview,
        }
        import_session_id = create_import_session(
            conn,
            user_id=user_id,
            filename=filename,
            file_size=file_size,
            payload=payload,
        )
        conn.close()
        return jsonify({
            'status': 'review_required',
            'message': (
                'Duplicates found. Review and confirm before importing.'
                if source == 'aidiary'
                else 'Review and confirm entries before importing.'
            ),
            'summary': preview['summary'],
            'duplicate_entries': public_duplicate_rows,
            'review_entries': public_review_rows,
            'warnings': parse_warnings,
            'errors': [],
            'import_session_id': import_session_id,
            'import_id': None,
        }), 200

    import_id = create_pending_import_history(
        conn,
        user_id=user_id,
        filename=filename,
        file_size=file_size,
    )
    result = commit_import_preview(conn, user_id, preview, import_id, set())
    all_warnings = parse_warnings[:]
    any_inserted = _import_inserted_total(result) > 0
    status_str = 'success' if any_inserted else 'skipped'
    finalise_import_history(
        conn,
        import_id=import_id,
        user_id=user_id,
        result=result,
        warnings=all_warnings,
        status=status_str,
    )
    conn.close()

    return jsonify({
        'status': status_str,
        'summary': {
            'inserted_daily': result['inserted_daily'],
            'skipped_daily': result['skipped_daily'],
            'inserted_dreams': result['inserted_dreams'],
            'skipped_dreams': result['skipped_dreams'],
            'inserted_important_days': result.get('inserted_important_days', 0),
            'inserted_thought_records': result.get('inserted_thought_records', 0),
            'duplicate_dates_daily': result['duplicate_dates_daily'],
            'duplicate_dates_dreams': result['duplicate_dates_dreams'],
        },
        'duplicate_entries': [],
        'warnings': all_warnings,
        'errors': [],
        'import_id': import_id,
        'import_session_id': None,
    }), 200


@import_bp.route('/import/commit', methods=['POST'])
@jwt_required()
def commit_import():
    """Commit a staged import session after duplicate review."""
    user_id = int(get_jwt_identity())
    payload = request.get_json(silent=True) or {}
    import_session_id = payload.get('import_session_id')
    accepted_duplicate_row_ids = payload.get('accepted_duplicate_row_ids', [])
    selected_row_ids = payload.get('selected_row_ids')
    entry_type_overrides = payload.get('entry_type_overrides', {})

    if not isinstance(import_session_id, str) or not import_session_id.strip():
        return jsonify({'status': 'error', 'errors': ['Import session id is required.']}), 400

    if not isinstance(accepted_duplicate_row_ids, list) or any(
        not isinstance(row_id, str) for row_id in accepted_duplicate_row_ids
    ):
        return jsonify({
            'status': 'error',
            'errors': ['Accepted duplicate row ids must be an array of strings.'],
        }), 400
    if selected_row_ids is not None and (
        not isinstance(selected_row_ids, list)
        or any(not isinstance(row_id, str) for row_id in selected_row_ids)
    ):
        return jsonify({'status': 'error', 'errors': ['Selected row ids must be an array of strings.']}), 400
    if not isinstance(entry_type_overrides, dict) or any(
        not isinstance(key, str) or value not in {'daily', 'dream'}
        for key, value in entry_type_overrides.items()
    ):
        return jsonify({'status': 'error', 'errors': ['Entry type overrides are invalid.']}), 400

    conn = get_db()
    ensure_history_table(conn)
    ensure_import_sessions_table(conn)
    session = get_import_session(conn, user_id=user_id, session_id=import_session_id)
    if not session:
        conn.close()
        return jsonify({
            'status': 'error',
            'errors': ['The import review session could not be found or has already been used.'],
        }), 404

    preview_payload = session.get('payload', {})
    parse_warnings = preview_payload.get('parse_warnings', [])

    try:
        import_id = create_pending_import_history(
            conn,
            user_id=user_id,
            filename=session['filename'],
            file_size=session['file_size_bytes'],
        )
        result = commit_import_preview(
            conn,
            user_id,
            preview_payload,
            import_id,
            set(accepted_duplicate_row_ids),
            set(selected_row_ids) if selected_row_ids is not None else None,
            entry_type_overrides,
        )
        status_str = 'success'
        if result['skipped_daily'] or result['skipped_dreams']:
            status_str = 'partial' if _import_inserted_total(result) > 0 else 'skipped'

        final_warnings = list(parse_warnings)
        omitted_duplicates = result['skipped_daily'] + result['skipped_dreams']
        if omitted_duplicates:
            final_warnings.append(
                f'{omitted_duplicates} duplicate entr{"y" if omitted_duplicates == 1 else "ies"} were left out.'
            )

        finalise_import_history(
            conn,
            import_id=import_id,
            user_id=user_id,
            result=result,
            warnings=final_warnings,
            status=status_str,
        )
        mark_import_session_consumed(conn, import_session_id)
    except Exception as exc:
        conn.close()
        current_app.logger.error('Import commit failed: %s', exc)
        return jsonify({'status': 'error', 'errors': ['Database error during import commit.']}), 500

    conn.close()
    return jsonify({
        'status': status_str,
        'message': 'Import successful.',
        'summary': {
            'inserted_daily': result['inserted_daily'],
            'skipped_daily': result['skipped_daily'],
            'inserted_dreams': result['inserted_dreams'],
            'skipped_dreams': result['skipped_dreams'],
            'inserted_important_days': result.get('inserted_important_days', 0),
            'inserted_thought_records': result.get('inserted_thought_records', 0),
            'duplicate_dates_daily': result['duplicate_dates_daily'],
            'duplicate_dates_dreams': result['duplicate_dates_dreams'],
        },
        'duplicate_entries': [],
        'warnings': final_warnings,
        'errors': [],
        'import_id': import_id,
        'import_session_id': None,
    }), 200


def _run_import_job(
    app,
    *,
    job_id: str,
    user_id: int,
    import_session_id: str,
    accepted_duplicate_row_ids: set[str],
    selected_row_ids: set[str] | None,
    entry_type_overrides: dict[str, str],
) -> None:
    """Run a reviewed import outside the request lifecycle."""
    conn = None
    import_id = None
    try:
        with app.app_context():
            conn = get_db()
            ensure_history_table(conn)
            ensure_import_sessions_table(conn)
            ensure_import_jobs_table(conn)
            job_row = conn.execute(
                'SELECT import_id FROM import_jobs WHERE id = ? AND user_id = ?',
                (job_id, user_id),
            ).fetchone()
            if not job_row:
                raise ValueError('The durable import job could not be found.')
            session = get_import_session(
                conn,
                user_id=user_id,
                session_id=import_session_id,
            )
            if not session:
                completed_history = None
                if job_row['import_id'] is not None:
                    completed_history = conn.execute(
                        '''SELECT inserted_daily, skipped_daily, inserted_dreams,
                                  skipped_dreams, warnings, status
                           FROM import_history
                           WHERE id = ? AND user_id = ? AND status != 'processing' ''',
                        (job_row['import_id'], user_id),
                    ).fetchone()
                if completed_history:
                    recovered_result = {
                        'inserted_daily': completed_history['inserted_daily'],
                        'skipped_daily': completed_history['skipped_daily'],
                        'inserted_dreams': completed_history['inserted_dreams'],
                        'skipped_dreams': completed_history['skipped_dreams'],
                        'duplicate_dates_daily': [],
                        'duplicate_dates_dreams': [],
                    }
                    try:
                        recovered_warnings = json.loads(completed_history['warnings'] or '[]')
                    except (TypeError, ValueError):
                        recovered_warnings = []
                    response = _build_commit_response(
                        recovered_result,
                        status=completed_history['status'],
                        warnings=recovered_warnings,
                        import_id=job_row['import_id'],
                    )
                    recovered_total = (
                        recovered_result['inserted_daily']
                        + recovered_result['inserted_dreams']
                        + recovered_result.get('inserted_important_days', 0)
                        + recovered_result.get('inserted_thought_records', 0)
                    )
                    _persist_import_job(
                        conn,
                        job_id,
                        status='completed',
                        processed=recovered_total,
                        total=recovered_total,
                        percent=100,
                        message=f'Import complete: {recovered_total} entries imported.',
                        result_json=json.dumps(response),
                        error=None,
                        completed_at=datetime.now(timezone.utc).strftime('%Y-%m-%dT%H:%M:%SZ'),
                    )
                    return
                raise ValueError('The import review session is no longer available.')

            preview_payload = session.get('payload', {})
            parse_warnings = list(preview_payload.get('parse_warnings', []))
            import_id = job_row['import_id']
            if import_id is None:
                import_id = create_pending_import_history(
                    conn,
                    user_id=user_id,
                    filename=session['filename'],
                    file_size=session['file_size_bytes'],
                )
            _persist_import_job(
                conn,
                job_id,
                status='running',
                import_id=import_id,
                started_at=datetime.now(timezone.utc).strftime('%Y-%m-%dT%H:%M:%SZ'),
                attempt_count=conn.execute(
                    'SELECT attempt_count FROM import_jobs WHERE id = ?',
                    (job_id,),
                ).fetchone()[0] + 1,
                message='Preparing selected entries…',
            )
            _update_import_job(job_id, status='running', message='Preparing selected entries…')

            def update_progress(processed: int, total: int) -> None:
                _update_import_job(
                    job_id,
                    processed=processed,
                    total=total,
                    percent=round((processed / total) * 100) if total else 100,
                    message=(
                        f'Importing {processed} of {total} entries…'
                        if processed < total
                        else 'Finalising import…'
                    ),
                )

            result = commit_import_preview(
                conn,
                user_id,
                preview_payload,
                import_id,
                accepted_duplicate_row_ids,
                selected_row_ids,
                entry_type_overrides,
                progress_callback=update_progress,
            )
            status_str = 'success'
            if result['skipped_daily'] or result['skipped_dreams']:
                status_str = (
                    'partial'
                    if _import_inserted_total(result) > 0
                    else 'skipped'
                )

            omitted_duplicates = result['skipped_daily'] + result['skipped_dreams']
            if omitted_duplicates:
                parse_warnings.append(
                    f'{omitted_duplicates} duplicate '
                    f'entr{"y" if omitted_duplicates == 1 else "ies"} were left out.'
                )

            finalise_import_history(
                conn,
                import_id=import_id,
                user_id=user_id,
                result=result,
                warnings=parse_warnings,
                status=status_str,
            )
            mark_import_session_consumed(conn, import_session_id)
            response = _build_commit_response(
                result,
                status=status_str,
                warnings=parse_warnings,
                import_id=import_id,
            )
            _update_import_job(
                job_id,
                status='completed',
                percent=100,
                message=(
                    f'Import complete: {_import_inserted_total(result)} records imported.'
                ),
                result=response,
            )
            _persist_import_job(
                conn,
                job_id,
                status='completed',
                processed=conn.execute(
                    'SELECT total FROM import_jobs WHERE id = ?',
                    (job_id,),
                ).fetchone()[0],
                percent=100,
                message=(
                    f'Import complete: {_import_inserted_total(result)} records imported.'
                ),
                result_json=json.dumps(response),
                error=None,
                worker_token=None,
                lease_expires_at=None,
                completed_at=datetime.now(timezone.utc).strftime('%Y-%m-%dT%H:%M:%SZ'),
            )
    except Exception as exc:
        if conn:
            conn.rollback()
        app.logger.error('Background import job %s failed: %s', job_id, exc)
        _update_import_job(
            job_id,
            status='failed',
            message='Import failed. Return to Import for details or try the upload again.',
            error=str(exc) if isinstance(exc, ValueError) else 'Database error during import.',
        )
        try:
            with app.app_context():
                failure_conn = get_db()
                ensure_import_jobs_table(failure_conn)
                _persist_import_job(
                    failure_conn,
                    job_id,
                    status='failed',
                    message='Import failed. Return to Import for details or try the upload again.',
                    error=(
                        str(exc)
                        if isinstance(exc, ValueError)
                        else 'Database error during import.'
                    ),
                    worker_token=None,
                    lease_expires_at=None,
                    completed_at=datetime.now(timezone.utc).strftime('%Y-%m-%dT%H:%M:%SZ'),
                )
                failure_conn.close()
        except Exception as persist_exc:
            app.logger.error('Could not persist import job failure %s: %s', job_id, persist_exc)
    finally:
        if conn:
            conn.close()
        with _IMPORT_JOBS_LOCK:
            _ACTIVE_IMPORT_JOB_THREADS.discard(job_id)
            _IMPORT_JOB_PROGRESS.pop(job_id, None)


def _launch_import_job(app, job_id: str) -> bool:
    """Launch or recover one durable job once within this server process."""
    with _IMPORT_JOBS_LOCK:
        if job_id in _ACTIVE_IMPORT_JOB_THREADS:
            return False
        _ACTIVE_IMPORT_JOB_THREADS.add(job_id)

    try:
        with app.app_context():
            conn = get_db()
            ensure_import_jobs_table(conn)
            now = datetime.now(timezone.utc)
            now_text = now.strftime('%Y-%m-%dT%H:%M:%SZ')
            worker_token = secrets.token_urlsafe(18)
            lease_expires_at = (now + timedelta(minutes=30)).strftime('%Y-%m-%dT%H:%M:%SZ')
            claimed = conn.execute(
                '''UPDATE import_jobs
                   SET status = 'running', worker_token = ?, lease_expires_at = ?,
                       updated_at = ?
                   WHERE id = ?
                     AND (
                       status = 'queued'
                       OR (
                         status = 'running'
                         AND (lease_expires_at IS NULL OR lease_expires_at < ?)
                       )
                     )''',
                (worker_token, lease_expires_at, now_text, job_id, now_text),
            )
            conn.commit()
            if claimed.rowcount != 1:
                conn.close()
                with _IMPORT_JOBS_LOCK:
                    _ACTIVE_IMPORT_JOB_THREADS.discard(job_id)
                return False
            row = conn.execute(
                '''SELECT user_id, import_session_id, request_json, status
                   FROM import_jobs WHERE id = ?''',
                (job_id,),
            ).fetchone()
            conn.close()
        if not row or row['status'] not in {'queued', 'running'}:
            with _IMPORT_JOBS_LOCK:
                _ACTIVE_IMPORT_JOB_THREADS.discard(job_id)
            return False
        request_payload = json.loads(row['request_json'] or '{}')
    except (TypeError, ValueError, sqlite3.Error):
        with _IMPORT_JOBS_LOCK:
            _ACTIVE_IMPORT_JOB_THREADS.discard(job_id)
        return False

    thread = threading.Thread(
        target=_run_import_job,
        kwargs={
            'app': app,
            'job_id': job_id,
            'user_id': int(row['user_id']),
            'import_session_id': row['import_session_id'],
            'accepted_duplicate_row_ids': set(
                request_payload.get('accepted_duplicate_row_ids', [])
            ),
            'selected_row_ids': (
                set(request_payload['selected_row_ids'])
                if request_payload.get('selected_row_ids') is not None
                else None
            ),
            'entry_type_overrides': request_payload.get('entry_type_overrides', {}),
        },
        daemon=True,
        name=f'import-{job_id[:8]}',
    )
    thread.start()
    return True


def recover_import_jobs(app) -> int:
    """Restart durable queued/running jobs after application startup."""
    with app.app_context():
        conn = get_db()
        ensure_import_jobs_table(conn)
        now = datetime.now(timezone.utc).strftime('%Y-%m-%dT%H:%M:%SZ')
        rows = conn.execute(
            '''SELECT id FROM import_jobs
               WHERE status = 'queued'
                  OR (
                    status = 'running'
                    AND (lease_expires_at IS NULL OR lease_expires_at < ?)
                  )
               ORDER BY created_at ASC''',
            (now,),
        ).fetchall()
        conn.close()

    recovered = 0
    for row in rows:
        if _launch_import_job(app, row['id']):
            recovered += 1
    return recovered


@import_bp.route('/import/jobs', methods=['POST'])
@jwt_required()
def start_import_job():
    """Start a reviewed import and return immediately so the UI can navigate away."""
    user_id = int(get_jwt_identity())
    payload = request.get_json(silent=True) or {}
    import_session_id = payload.get('import_session_id')
    accepted_duplicate_row_ids = payload.get('accepted_duplicate_row_ids', [])
    selected_row_ids = payload.get('selected_row_ids')
    entry_type_overrides = payload.get('entry_type_overrides', {})

    if not isinstance(import_session_id, str) or not import_session_id.strip():
        return jsonify({'status': 'error', 'errors': ['Import session id is required.']}), 400
    if not isinstance(accepted_duplicate_row_ids, list) or any(
        not isinstance(row_id, str) for row_id in accepted_duplicate_row_ids
    ):
        return jsonify({'status': 'error', 'errors': ['Accepted duplicate row ids are invalid.']}), 400
    if selected_row_ids is not None and (
        not isinstance(selected_row_ids, list)
        or any(not isinstance(row_id, str) for row_id in selected_row_ids)
    ):
        return jsonify({'status': 'error', 'errors': ['Selected row ids are invalid.']}), 400
    if not isinstance(entry_type_overrides, dict) or any(
        not isinstance(key, str) or value not in {'daily', 'dream'}
        for key, value in entry_type_overrides.items()
    ):
        return jsonify({'status': 'error', 'errors': ['Entry type overrides are invalid.']}), 400

    conn = get_db()
    ensure_import_jobs_table(conn)
    session = get_import_session(
        conn,
        user_id=user_id,
        session_id=import_session_id,
    )
    if not session:
        conn.close()
        return jsonify({'status': 'error', 'errors': ['Import review session not found.']}), 404
    active_job = conn.execute(
        '''SELECT id FROM import_jobs
           WHERE user_id = ? AND import_session_id = ? AND status IN ('queued', 'running')
           LIMIT 1''',
        (user_id, import_session_id),
    ).fetchone()
    if active_job:
        conn.close()
        return jsonify({'status': 'error', 'errors': ['This import is already running.']}), 409

    job_id = secrets.token_urlsafe(18)
    now = datetime.now(timezone.utc).strftime('%Y-%m-%dT%H:%M:%SZ')
    request_payload = {
        'accepted_duplicate_row_ids': accepted_duplicate_row_ids,
        'selected_row_ids': selected_row_ids,
        'entry_type_overrides': entry_type_overrides,
    }
    conn.execute(
        '''INSERT INTO import_jobs
           (id, user_id, import_session_id, status, processed, total, percent,
            message, request_json, created_at, updated_at)
           VALUES (?, ?, ?, 'queued', 0, ?, 0, ?, ?, ?, ?)''',
        (
            job_id,
            user_id,
            import_session_id,
            len(selected_row_ids or []),
            'Import queued…',
            json.dumps(request_payload),
            now,
            now,
        ),
    )
    conn.commit()
    conn.close()

    _launch_import_job(current_app._get_current_object(), job_id)
    return jsonify(_get_import_job(job_id, user_id)), 202


@import_bp.route('/import/jobs/<job_id>', methods=['GET'])
@jwt_required()
def get_import_job_status(job_id: str):
    user_id = int(get_jwt_identity())
    job = _get_import_job(job_id, user_id)
    if not job:
        return jsonify({'status': 'error', 'errors': ['Import job not found.']}), 404
    if job['status'] in {'queued', 'running'}:
        _launch_import_job(current_app._get_current_object(), job_id)
    return jsonify(job), 200


# ---------------------------------------------------------------------------
# GET /api/import/history
# ---------------------------------------------------------------------------

@import_bp.route('/import/history', methods=['GET'])
@jwt_required()
def get_history():
    """
    Return the import history for the authenticated user.

    Response 200:
    {
      "history": [
        {
          "id": int,
          "imported_at": "ISO-8601",
          "filename": str,
          "file_size_bytes": int,
          "inserted_daily": int,
          "skipped_daily": int,
          "inserted_dreams": int,
          "skipped_dreams": int,
          "warnings": [str, ...],
          "status": "success" | "skipped" | "empty"
        },
        ...
      ]
    }
    """
    user_id = int(get_jwt_identity())

    conn = get_db()
    ensure_history_table(conn)
    history = get_import_history(conn, user_id)
    conn.close()

    return jsonify({'history': history}), 200


@import_bp.route('/import/history/<int:import_id>/revert', methods=['POST'])
@jwt_required()
def revert_import(import_id: int):
    """Remove entries and stored media created by one completed import."""
    user_id = int(get_jwt_identity())
    data = request.get_json(silent=True) or {}
    if str(data.get('confirmation_text') or '').strip() != 'REVERT IMPORT':
        return jsonify({'error': 'Type REVERT IMPORT to confirm this action.'}), 400

    conn = get_db()
    ensure_history_table(conn)
    history = conn.execute(
        'SELECT id, status FROM import_history WHERE id = ? AND user_id = ?',
        (import_id, user_id),
    ).fetchone()
    if not history:
        conn.close()
        return jsonify({'error': 'Import history record not found.'}), 404
    if history['status'] == 'reverted':
        conn.close()
        return jsonify({'error': 'This import has already been reverted.'}), 409

    media_rows = conn.execute(
        '''SELECT image_storage_key AS storage_key FROM dailydiary_entries
           WHERE user_id = ? AND import_id = ? AND image_storage_key IS NOT NULL
           UNION ALL
           SELECT image_storage_key AS storage_key FROM dreamdiary_entries
           WHERE user_id = ? AND import_id = ? AND image_storage_key IS NOT NULL
           UNION ALL
           SELECT asset.storage_key FROM entry_assets asset
           JOIN dailydiary_entries entry ON entry.id = asset.entry_id
           WHERE asset.user_id = ? AND asset.entry_type = 'daily' AND entry.import_id = ?
           UNION ALL
           SELECT asset.storage_key FROM entry_assets asset
           JOIN dreamdiary_entries entry ON entry.id = asset.entry_id
           WHERE asset.user_id = ? AND asset.entry_type = 'dream' AND entry.import_id = ?''',
        (user_id, import_id, user_id, import_id, user_id, import_id, user_id, import_id),
    ).fetchall()
    storage_keys = [str(row['storage_key']) for row in media_rows if row['storage_key']]
    try:
        conn.execute(
            '''DELETE FROM entry_assets WHERE user_id = ? AND (
               (entry_type = 'daily' AND entry_id IN (
                   SELECT id FROM dailydiary_entries WHERE user_id = ? AND import_id = ?
               )) OR (entry_type = 'dream' AND entry_id IN (
                   SELECT id FROM dreamdiary_entries WHERE user_id = ? AND import_id = ?
               )))''',
            (user_id, user_id, import_id, user_id, import_id),
        )
        daily_deleted = conn.execute(
            'DELETE FROM dailydiary_entries WHERE user_id = ? AND import_id = ?',
            (user_id, import_id),
        ).rowcount
        dream_deleted = conn.execute(
            'DELETE FROM dreamdiary_entries WHERE user_id = ? AND import_id = ?',
            (user_id, import_id),
        ).rowcount
        conn.execute(
            "UPDATE import_history SET status = 'reverted' WHERE id = ? AND user_id = ?",
            (import_id, user_id),
        )
        conn.commit()
    except Exception:
        conn.rollback()
        conn.close()
        raise
    conn.close()

    for storage_key in storage_keys:
        delete_image(storage_key)
    return jsonify({
        'message': 'Import reverted.',
        'deleted_daily': daily_deleted,
        'deleted_dreams': dream_deleted,
        'deleted_total': daily_deleted + dream_deleted,
    }), 200


# ---------------------------------------------------------------------------
# GET /api/import/template
# ---------------------------------------------------------------------------

@import_bp.route('/import/template', methods=['GET'])
@jwt_required()
def download_template():
    """
    Generate and return a blank Excel import template (.xlsx) with the
    correct sheet names and column headers for Daily and Dreams entries.
    """
    try:
        import openpyxl
    except ImportError:
        return jsonify({
            'status': 'error',
            'errors': ['openpyxl is not installed on the server.'],
        }), 500

    wb = openpyxl.Workbook()

    # --- Daily sheet ---
    ws_daily = wb.active
    ws_daily.title = 'Daily'
    ws_daily.append(list(DAILY_IMPORT_HEADERS))

    # --- Dreams sheet ---
    ws_dreams = wb.create_sheet(title='Dreams')
    ws_dreams.append(list(DREAM_IMPORT_HEADERS))

    ws_important_days = wb.create_sheet(title='Important Days')
    ws_important_days.append(list(IMPORTANT_DAY_IMPORT_HEADERS))

    ws_thought_records = wb.create_sheet(title='Thought Records')
    ws_thought_records.append(list(THOUGHT_RECORD_IMPORT_HEADERS))

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return send_file(
        buffer,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='openmynd_import_template.xlsx',
    )


# ---------------------------------------------------------------------------
# GET /api/import/export
# ---------------------------------------------------------------------------

@import_bp.route('/import/export', methods=['GET'])
@jwt_required()
def export_entries():
    """Export the authenticated user's entries into an Excel workbook."""
    user_id = int(get_jwt_identity())

    from_date_raw = request.args.get('from_date')
    to_date_raw = request.args.get('to_date')

    include_daily_raw = request.args.get('include_daily', 'true')
    include_dreams_raw = request.args.get('include_dreams', 'true')
    include_important_days_raw = request.args.get('include_important_days', 'false')
    include_thought_records_raw = request.args.get('include_thought_records', 'false')
    export_all = str(request.args.get('export_all', 'false')).strip().lower() == 'true'

    include_daily = str(include_daily_raw).strip().lower() != 'false'
    include_dreams = str(include_dreams_raw).strip().lower() != 'false'
    include_important_days = str(include_important_days_raw).strip().lower() == 'true'
    include_thought_records = str(include_thought_records_raw).strip().lower() == 'true'

    if export_all:
        include_daily = True
        include_dreams = True
        include_important_days = True
        include_thought_records = True
        from_date_raw = None
        to_date_raw = None

    if not include_daily and not include_dreams and not include_important_days and not include_thought_records:
        return jsonify({
            'status': 'error',
            'errors': ['At least one export type must be selected.'],
        }), 400

    from_date = None
    to_date = None

    if from_date_raw:
        try:
            from_date = datetime.strptime(from_date_raw, '%Y-%m-%d').date()
        except ValueError:
            return jsonify({
                'status': 'error',
                'errors': ['from_date must be in YYYY-MM-DD format.'],
            }), 400

    if to_date_raw:
        try:
            to_date = datetime.strptime(to_date_raw, '%Y-%m-%d').date()
        except ValueError:
            return jsonify({
                'status': 'error',
                'errors': ['to_date must be in YYYY-MM-DD format.'],
            }), 400

    if from_date and to_date and from_date > to_date:
        return jsonify({
            'status': 'error',
            'errors': ['from_date cannot be after to_date.'],
        }), 400

    try:
        import openpyxl
    except ImportError:
        return jsonify({
            'status': 'error',
            'errors': ['openpyxl is not installed on the server.'],
        }), 500

    conn = get_db()

    daily_rows = []
    dream_rows = []
    important_day_rows = []
    thought_record_rows = []

    first_daily = conn.execute(
        'SELECT MIN(entry_date) AS min_date, COUNT(*) AS total_count FROM dailydiary_entries WHERE user_id = ?',
        (user_id,),
    ).fetchone()
    first_dream = conn.execute(
        'SELECT MIN(entry_date) AS min_date, COUNT(*) AS total_count FROM dreamdiary_entries WHERE user_id = ?',
        (user_id,),
    ).fetchone()

    overall_first_date = min(
        [value for value in [first_daily['min_date'], first_dream['min_date']] if value],
        default=None,
    )
    overall_last_daily = conn.execute(
        'SELECT MAX(entry_date) AS max_date FROM dailydiary_entries WHERE user_id = ?',
        (user_id,),
    ).fetchone()
    overall_last_dream = conn.execute(
        'SELECT MAX(entry_date) AS max_date FROM dreamdiary_entries WHERE user_id = ?',
        (user_id,),
    ).fetchone()
    overall_last_date = max(
        [value for value in [overall_last_daily['max_date'], overall_last_dream['max_date']] if value],
        default=None,
    )

    if include_daily:
        daily_query = '''
            SELECT id, entry_date, entry_time, entry_number, title, user_message, ai_response,
                   image_storage_key, image_source, image_position_x, image_position_y,
                   image_prompt, recycled_image_prompt
            FROM dailydiary_entries
            WHERE user_id = ?
        '''
        daily_params = [user_id]

        if from_date:
            daily_query += ' AND entry_date >= ?'
            daily_params.append(from_date.isoformat())

        if to_date:
            daily_query += ' AND entry_date <= ?'
            daily_params.append(to_date.isoformat())

        daily_query += " ORDER BY entry_date ASC, COALESCE(entry_time, '19:00') ASC, entry_number ASC"
        daily_rows = conn.execute(daily_query, tuple(daily_params)).fetchall()

    if include_dreams:
        dream_query = '''
            SELECT id, entry_date, entry_time, entry_number, title, plot, "cast", location, period,
                   emotion, symbols_and_imagery, insight, action, other, tags,
                   image_storage_key, image_source, image_position_x, image_position_y,
                   image_prompt, recycled_image_prompt
            FROM dreamdiary_entries
            WHERE user_id = ?
        '''
        dream_params = [user_id]

        if from_date:
            dream_query += ' AND entry_date >= ?'
            dream_params.append(from_date.isoformat())

        if to_date:
            dream_query += ' AND entry_date <= ?'
            dream_params.append(to_date.isoformat())

        dream_query += " ORDER BY entry_date ASC, COALESCE(entry_time, '08:00') ASC, entry_number ASC"
        dream_rows = conn.execute(dream_query, tuple(dream_params)).fetchall()

    if include_important_days:
        important_day_query = '''
            SELECT id, starts_on, label, category, recurrence, icon_name, accent_color,
                   note, image_storage_key
            FROM important_days
            WHERE user_id = ?
        '''
        important_day_params = [user_id]
        if from_date:
            important_day_query += ' AND starts_on >= ?'
            important_day_params.append(from_date.isoformat())
        if to_date:
            important_day_query += ' AND starts_on <= ?'
            important_day_params.append(to_date.isoformat())
        important_day_query += ' ORDER BY starts_on ASC, lower(label) ASC, id ASC'
        important_day_rows = conn.execute(important_day_query, tuple(important_day_params)).fetchall()

    if include_thought_records:
        thought_record_query = '''
            SELECT w.id, w.record_date, w.title, w.status, w.current_step,
                   w.linked_entry_type, w.linked_entry_id,
                   d.situation, d.feelings_before_json, d.unhelpful_thoughts,
                   d.evidence_for, d.evidence_against, d.balanced_thought,
                   d.feelings_after_json, d.next_step, d.ai_response,
                   d.ai_responded_at, d.ai_response_outdated
            FROM cbt_worksheets w
            JOIN cbt_thought_record_data d ON d.worksheet_id = w.id
            WHERE w.user_id = ?
        '''
        thought_record_params = [user_id]
        if from_date:
            thought_record_query += ' AND w.record_date >= ?'
            thought_record_params.append(from_date.isoformat())
        if to_date:
            thought_record_query += ' AND w.record_date <= ?'
            thought_record_params.append(to_date.isoformat())
        thought_record_query += ' ORDER BY w.record_date ASC, w.id ASC'
        thought_record_rows = conn.execute(thought_record_query, tuple(thought_record_params)).fetchall()

    wb = openpyxl.Workbook()
    daily_attachment_rows = _load_attachment_export_rows(
        conn,
        user_id=user_id,
        entry_type='daily',
        entry_ids=[int(row['id']) for row in daily_rows],
    ) if include_daily else {}
    dream_attachment_rows = _load_attachment_export_rows(
        conn,
        user_id=user_id,
        entry_type='dream',
        entry_ids=[int(row['id']) for row in dream_rows],
    ) if include_dreams else {}

    workbook_has_sheet = False

    def next_sheet(title: str):
        nonlocal workbook_has_sheet
        if not workbook_has_sheet:
            sheet = wb.active
            sheet.title = title
            workbook_has_sheet = True
            return sheet
        return wb.create_sheet(title=title)

    ws_daily = None
    ws_dreams = None
    ws_important_days = None

    if include_daily:
        ws_daily = next_sheet('Daily')
        ws_daily.title = 'Daily'
        ws_daily.append(list(DAILY_IMPORT_HEADERS))
        for row in daily_rows:
            ws_daily.append([
                row['entry_date'] or '',
                row['entry_time'] or '',
                row['title'] or '',
                row['user_message'] or '',
                row['ai_response'] or '',
                '',
            ])

    if include_dreams:
        ws_dreams = next_sheet('Dreams')
        ws_dreams.append(list(DREAM_IMPORT_HEADERS))
        for row in dream_rows:
            ws_dreams.append([
                row['entry_date'] or '',
                row['entry_time'] or '',
                row['title'] or '',
                row['plot'] or '',
                row['cast'] or '',
                row['location'] or '',
                row['period'] or '',
                row['emotion'] or '',
                row['symbols_and_imagery'] or '',
                row['insight'] or '',
                row['action'] or '',
                row['other'] or '',
                row['tags'] or '',
                '',
            ])

    if include_important_days:
        ws_important_days = next_sheet('Important Days')
        ws_important_days.append(list(IMPORTANT_DAY_IMPORT_HEADERS))
        for row in important_day_rows:
            ws_important_days.append([
                row['starts_on'] or '',
                row['label'] or '',
                row['category'] or '',
                row['recurrence'] or '',
                row['icon_name'] or '',
                row['accent_color'] or '',
                row['note'] or '',
                '',
            ])

    if include_thought_records:
        ws_thought_records = next_sheet('Thought Records')
        ws_thought_records.append(list(THOUGHT_RECORD_IMPORT_HEADERS))
        for row in thought_record_rows:
            ws_thought_records.append([
                row['record_date'] or '',
                row['title'] or '',
                row['status'] or '',
                row['current_step'] or '',
                row['linked_entry_type'] or '',
                row['linked_entry_id'] or '',
                row['situation'] or '',
                row['feelings_before_json'] or '[]',
                row['unhelpful_thoughts'] or '',
                row['evidence_for'] or '',
                row['evidence_against'] or '',
                row['balanced_thought'] or '',
                row['feelings_after_json'] or '[]',
                row['next_step'] or '',
                row['ai_response'] or '',
                row['ai_responded_at'] or '',
                row['ai_response_outdated'] or 0,
            ])

    stamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    has_filters = (
        not export_all
        and (
            from_date is not None
            or to_date is not None
            or not include_daily
            or not include_dreams
            or include_important_days
            or include_thought_records
        )
    )
    filename = f'openmynd_export_{stamp}.zip'
    if has_filters:
        filename = f'openmynd_export_filtered_{stamp}.zip'

    ensure_export_history_table(conn)
    is_full_range = bool(
        include_daily
        and include_dreams
        and overall_first_date
        and overall_last_date
        and from_date
        and to_date
        and from_date.isoformat() == overall_first_date
        and to_date.isoformat() == overall_last_date
    )
    export_record = record_export_history(
        conn,
        user_id=user_id,
        filename=filename,
        from_date=from_date.isoformat() if from_date else None,
        to_date=to_date.isoformat() if to_date else None,
        include_daily=include_daily,
        include_dreams=include_dreams,
        daily_count=len(daily_rows),
        dream_count=len(dream_rows),
        is_full_range=is_full_range,
        issue_guard_token=is_full_range and (len(daily_rows) + len(dream_rows) > 0),
    )
    conn.close()

    manifest_assets: dict[str, dict[str, object]] = {}
    if include_daily:
        for row_index, row in enumerate(daily_rows, start=2):
            storage_key = row['image_storage_key']
            entry_attachments = daily_attachment_rows.get(int(row['id']), [])
            if not storage_key and not entry_attachments:
                continue
            image_bytes = read_media_bytes(storage_key)
            if storage_key and not image_bytes:
                continue
            asset_ref = _build_entry_asset_ref('daily', row)
            ws_daily.cell(row=row_index, column=len(DAILY_IMPORT_HEADERS)).value = asset_ref
            image_filename = _image_filename_for_storage_key(storage_key)
            manifest_assets[asset_ref] = {
                'entry_type': 'daily',
                'image_filename': image_filename,
                'image_source': row['image_source'],
                'image_position_x': row['image_position_x'],
                'image_position_y': row['image_position_y'],
                'image_prompt': row['image_prompt'],
                'recycled_image_prompt': row['recycled_image_prompt'],
                'attachments': [
                    {
                        'package_filename': _package_attachment_filename(attachment, index),
                        'original_filename': attachment['original_filename'],
                        'mime_type': attachment['mime_type'],
                        'asset_role': attachment['asset_role'],
                        'file_size_bytes': int(attachment['file_size_bytes'] or 0),
                        'sort_order': int(attachment['sort_order'] or 0),
                    }
                    for index, attachment in enumerate(entry_attachments, start=1)
                    if attachment['storage_key'] and read_media_bytes(attachment['storage_key'])
                ],
            }

    if include_dreams:
        ws_target = ws_dreams
        for row_index, row in enumerate(dream_rows, start=2):
            storage_key = row['image_storage_key']
            entry_attachments = dream_attachment_rows.get(int(row['id']), [])
            if not storage_key and not entry_attachments:
                continue
            image_bytes = read_media_bytes(storage_key)
            if storage_key and not image_bytes:
                continue
            asset_ref = _build_entry_asset_ref('dream', row)
            ws_target.cell(row=row_index, column=len(DREAM_IMPORT_HEADERS)).value = asset_ref
            image_filename = _image_filename_for_storage_key(storage_key)
            manifest_assets[asset_ref] = {
                'entry_type': 'dream',
                'image_filename': image_filename,
                'image_source': row['image_source'],
                'image_position_x': row['image_position_x'],
                'image_position_y': row['image_position_y'],
                'image_prompt': row['image_prompt'],
                'recycled_image_prompt': row['recycled_image_prompt'],
                'attachments': [
                    {
                        'package_filename': _package_attachment_filename(attachment, index),
                        'original_filename': attachment['original_filename'],
                        'mime_type': attachment['mime_type'],
                        'asset_role': attachment['asset_role'],
                        'file_size_bytes': int(attachment['file_size_bytes'] or 0),
                        'sort_order': int(attachment['sort_order'] or 0),
                    }
                    for index, attachment in enumerate(entry_attachments, start=1)
                    if attachment['storage_key'] and read_media_bytes(attachment['storage_key'])
                ],
            }

    if include_important_days and ws_important_days:
        for row_index, row in enumerate(important_day_rows, start=2):
            storage_key = row['image_storage_key']
            if not storage_key:
                continue
            image_bytes = read_media_bytes(storage_key)
            if not image_bytes:
                continue
            asset_ref = _build_important_day_asset_ref(row)
            ws_important_days.cell(row=row_index, column=len(IMPORTANT_DAY_IMPORT_HEADERS)).value = asset_ref
            manifest_assets[asset_ref] = {
                'entry_type': 'important_day',
                'image_filename': _image_filename_for_storage_key(storage_key),
                'image_source': 'upload',
                'image_position_x': None,
                'image_position_y': None,
                'image_prompt': None,
                'recycled_image_prompt': None,
                'attachments': [],
            }

    final_workbook_buffer = io.BytesIO()
    wb.save(final_workbook_buffer)
    final_workbook_buffer.seek(0)

    package_buffer = io.BytesIO()
    with zipfile.ZipFile(package_buffer, mode='w', compression=zipfile.ZIP_DEFLATED) as package_zip:
        package_zip.writestr('entries.xlsx', final_workbook_buffer.getvalue())
        for asset_ref, asset_meta in manifest_assets.items():
            image_filename = asset_meta['image_filename']
            if asset_meta['entry_type'] == 'important_day':
                source_rows = important_day_rows
                source_row = next(
                    (
                        row
                        for row in source_rows
                        if _build_important_day_asset_ref(row) == asset_ref
                    ),
                    None,
                )
            else:
                source_rows = daily_rows if asset_meta['entry_type'] == 'daily' else dream_rows
                source_row = next(
                    (
                        row
                        for row in source_rows
                        if _build_entry_asset_ref(asset_meta['entry_type'], row) == asset_ref
                    ),
                    None,
                )
            if source_row is None:
                continue

            if image_filename:
                image_bytes = read_media_bytes(source_row['image_storage_key'])
                if image_bytes:
                    package_zip.writestr(f'media/{asset_ref}/{image_filename}', image_bytes)

            for attachment_meta in asset_meta.get('attachments', []):
                if not isinstance(attachment_meta, dict):
                    continue
                package_filename = str(attachment_meta.get('package_filename') or '').strip()
                if not package_filename:
                    continue
                attachment_rows = (
                    daily_attachment_rows.get(int(source_row['id']), [])
                    if asset_meta['entry_type'] == 'daily'
                    else dream_attachment_rows.get(int(source_row['id']), [])
                )
                attachment_storage_key = next(
                    (
                        attachment_row['storage_key']
                        for index, attachment_row in enumerate(attachment_rows, start=1)
                        if _package_attachment_filename(attachment_row, index) == package_filename
                    ),
                    None,
                )
                attachment_bytes = read_media_bytes(attachment_storage_key)
                if not attachment_bytes:
                    continue
                package_zip.writestr(
                    f'media/{asset_ref}/{package_filename}',
                    attachment_bytes,
                )

        package_zip.writestr(
            'manifest.json',
            json.dumps(
                {
                    'package_type': PACKAGE_TYPE,
                    'version': PACKAGE_FORMAT_VERSION,
                    'generated_at': datetime.now(timezone.utc).strftime('%Y-%m-%dT%H:%M:%SZ'),
                    'portability': PORTABILITY_CONTRACT,
                    'assets': manifest_assets,
                },
                ensure_ascii=True,
                indent=2,
            ),
        )

    package_buffer.seek(0)

    response = send_file(
        package_buffer,
        mimetype='application/zip',
        as_attachment=True,
        download_name=filename,
    )
    response.headers['Access-Control-Expose-Headers'] = (
        'X-OpenMynd-Export-Token, X-AiDiary-Export-Token'
    )
    if export_record['guard_token']:
        response.headers['X-OpenMynd-Export-Token'] = export_record['guard_token']
        response.headers['X-AiDiary-Export-Token'] = export_record['guard_token']
    return response
