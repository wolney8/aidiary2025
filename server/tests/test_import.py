# server/tests/test_import.py
# Tests for import backend validation, duplicate handling, history, and API contract
import io
import base64
import json
import os
import sqlite3
import tempfile
import time
import zipfile
import shutil

import openpyxl
import pytest

from app import create_app
from services.import_service import (
    DAILY_IMPORT_HEADERS,
    DREAM_IMPORT_HEADERS,
    IMPORTANT_DAY_IMPORT_HEADERS,
    THOUGHT_RECORD_IMPORT_HEADERS,
    preview_import_entries,
)
from services.nltk_enrichment import derive_daily_nltk_fields


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

def _create_tables(conn: sqlite3.Connection) -> None:
    """Create all tables needed by the import tests."""
    conn.executescript('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY,
            username TEXT NOT NULL,
            password TEXT NOT NULL,
            first_name TEXT,
            last_name TEXT,
            age INTEGER,
            sex TEXT,
            goals TEXT,
            dailydiary_api_key TEXT,
            dreamdiary_api_key TEXT,
            chatgpt_daily_diary_coachname TEXT,
            chatgpt_dream_diary_coachname TEXT
        );

        CREATE TABLE IF NOT EXISTS dailydiary_entries (
            id INTEGER PRIMARY KEY,
            user_id INTEGER,
            import_id INTEGER,
            entry_date DATE,
            entry_time TEXT,
            entry_number INTEGER,
            title TEXT,
            user_message TEXT,
            ai_response TEXT,
            daily_people_names TEXT,
            daily_places TEXT,
            tags TEXT,
            mood TEXT,
            image_prompt TEXT,
            recycled_image_prompt TEXT,
            image_url TEXT,
            image_storage_key TEXT,
            image_position_x TEXT,
            image_position_y TEXT,
            image_source TEXT,
            FOREIGN KEY (user_id) REFERENCES users(id)
        );

        CREATE TABLE IF NOT EXISTS dreamdiary_entries (
            id INTEGER PRIMARY KEY,
            user_id INTEGER,
            import_id INTEGER,
            entry_date DATE,
            entry_time TEXT,
            entry_number INTEGER,
            title TEXT,
            cast TEXT,
            location TEXT,
            period TEXT,
            emotion TEXT,
            plot TEXT,
            symbols_and_imagery TEXT,
            insight TEXT,
            action TEXT,
            other TEXT,
            summary TEXT,
            interpretation TEXT,
            image_prompt TEXT,
            recycled_image_prompt TEXT,
            image_url TEXT,
            image_storage_key TEXT,
            image_position_x TEXT,
            image_position_y TEXT,
            image_source TEXT,
            dream_people_names TEXT,
            dream_places TEXT,
            tags TEXT,
            mood TEXT,
            FOREIGN KEY (user_id) REFERENCES users(id)
        );

        CREATE TABLE IF NOT EXISTS entry_assets (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            entry_type TEXT NOT NULL,
            entry_id INTEGER NOT NULL,
            asset_role TEXT NOT NULL DEFAULT 'attachment',
            storage_key TEXT NOT NULL,
            original_filename TEXT NOT NULL,
            mime_type TEXT NOT NULL,
            file_size_bytes INTEGER NOT NULL DEFAULT 0,
            sort_order INTEGER NOT NULL DEFAULT 0,
            created_at TEXT NOT NULL
        );
    ''')
    conn.commit()


class _PostgresDictRowCursor:
    """Small cursor double mirroring psycopg's configured dict_row result shape."""

    def __init__(self):
        self._rows: list[dict[str, str]] = []

    def execute(self, sql: str, _params=None):
        if "dailydiary_entries" in sql:
            self._rows = [
                {
                    "entry_date": "2026-08-23",
                    "entry_time": "09:00",
                    "title": "Existing daily",
                    "user_message": "Already stored.",
                }
            ]
        elif "dreamdiary_entries" in sql:
            self._rows = [
                {
                    "entry_date": "2026-08-22",
                    "entry_time": "08:00",
                    "title": "Existing dream",
                    "plot": "Already dreamed.",
                }
            ]
        return self

    def __iter__(self):
        return iter(self._rows)


class _PostgresDictRowConnection:
    def cursor(self):
        return _PostgresDictRowCursor()


def test_preview_import_supports_postgres_dict_rows_for_existing_entries():
    preview = preview_import_entries(
        _PostgresDictRowConnection(),
        user_id=7,
        parsed={
            "daily": [
                {
                    "entry_date": "2026-08-23",
                    "entry_time": "09:00",
                    "title": "Existing daily",
                    "user_message": "Already stored.",
                }
            ],
            "dreams": [
                {
                    "entry_date": "2026-08-22",
                    "entry_time": "08:00",
                    "title": "Existing dream",
                    "plot": "Already dreamed.",
                }
            ],
        },
    )

    assert preview["summary"]["duplicate_daily"] == 1
    assert preview["summary"]["duplicate_dreams"] == 1
    assert preview["ready_daily_rows"] == []
    assert preview["ready_dream_rows"] == []


@pytest.fixture
def client():
    """Flask test client with isolated in-memory database."""
    db_fd, db_path = tempfile.mkstemp(suffix='.db')
    media_root = tempfile.mkdtemp(prefix='aidiary-import-media-')
    os.environ['DB_PATH'] = db_path
    os.environ['JWT_SECRET'] = 'test-secret'
    os.environ['MEDIA_ROOT'] = media_root
    os.environ['AUTH_LOGIN_RATE_LIMIT'] = '1000 per minute'
    os.environ['AUTH_REGISTER_RATE_LIMIT'] = '1000 per minute'
    os.environ['IMPORT_UPLOAD_RATE_LIMIT'] = '1000 per minute'
    os.environ['IMPORT_COMMIT_RATE_LIMIT'] = '1000 per minute'
    os.environ['IMPORT_JOB_RATE_LIMIT'] = '1000 per minute'
    os.environ['IMPORT_REVERT_RATE_LIMIT'] = '1000 per minute'
    os.environ['EXPORT_RATE_LIMIT'] = '1000 per minute'

    app = create_app()
    app.config['TESTING'] = True

    with app.test_client() as test_client:
        conn = sqlite3.connect(db_path)
        _create_tables(conn)
        conn.close()
        yield test_client

    os.close(db_fd)
    os.unlink(db_path)
    shutil.rmtree(media_root, ignore_errors=True)


def _register_and_login(client) -> str:
    """Register a test user and return a JWT token."""
    client.post(
        '/api/register',
        data=json.dumps({'username': 'importer', 'password': 'secret123'}),
        content_type='application/json',
    )
    resp = client.post(
        '/api/login',
        data=json.dumps({'username': 'importer', 'password': 'secret123'}),
        content_type='application/json',
    )
    return json.loads(resp.data)['token']


def test_export_rate_limit_is_enforced(client, monkeypatch):
    monkeypatch.setenv('EXPORT_RATE_LIMIT', '1 per minute')
    token = _register_and_login(client)
    headers = {'Authorization': f'Bearer {token}'}
    url = '/api/import/export?include_daily=false&include_dreams=false'

    first = client.get(url, headers=headers)
    second = client.get(url, headers=headers)

    assert first.status_code == 400
    assert second.status_code == 429
    assert json.loads(second.data)['error'] == 'Too many attempts. Try again shortly.'


# ---------------------------------------------------------------------------
# Helpers to build Excel workbooks in memory
# ---------------------------------------------------------------------------

def _make_xlsx(daily_rows=None, dream_rows=None) -> bytes:
    """Create a minimal valid .xlsx workbook."""
    wb = openpyxl.Workbook()
    ws_daily = wb.active
    ws_daily.title = 'Daily'
    ws_daily.append(list(DAILY_IMPORT_HEADERS))
    for row in (daily_rows or []):
        row_values = list(row)
        if len(row_values) == 4:
            row_values = [row_values[0], '', *row_values[1:], '']
        ws_daily.append(row_values)

    ws_dreams = wb.create_sheet(title='Dreams')
    ws_dreams.append(list(DREAM_IMPORT_HEADERS))
    for row in (dream_rows or []):
        row_values = list(row)
        if len(row_values) == 12:
            row_values = [row_values[0], '', *row_values[1:], '']
        ws_dreams.append(row_values)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_xlsx_with_headers(
    daily_headers,
    dream_headers,
    daily_rows=None,
    dream_rows=None,
) -> bytes:
    """Create a workbook with explicit sheet headers for schema-contract tests."""
    wb = openpyxl.Workbook()
    ws_daily = wb.active
    ws_daily.title = 'Daily'
    ws_daily.append(daily_headers)
    for row in (daily_rows or []):
        row_values = list(row)
        if list(daily_headers) == list(DAILY_IMPORT_HEADERS) and len(row_values) == 4:
            row_values = [row_values[0], '', *row_values[1:], '']
        ws_daily.append(row_values)

    ws_dreams = wb.create_sheet(title='Dreams')
    ws_dreams.append(dream_headers)
    for row in (dream_rows or []):
        row_values = list(row)
        if list(dream_headers) == list(DREAM_IMPORT_HEADERS) and len(row_values) == 12:
            row_values = [row_values[0], '', *row_values[1:], '']
        ws_dreams.append(row_values)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _load_export_package(package_bytes: bytes):
    with zipfile.ZipFile(io.BytesIO(package_bytes)) as zf:
        workbook_bytes = zf.read('entries.xlsx')
        manifest = json.loads(zf.read('manifest.json').decode('utf-8'))
        workbook = openpyxl.load_workbook(io.BytesIO(workbook_bytes))
        return workbook, manifest, zf.namelist()


def _make_zip_package(
    daily_rows=None,
    dream_rows=None,
    manifest_assets=None,
    media_files=None,
    manifest_overrides=None,
) -> bytes:
    workbook_bytes = _make_xlsx_with_headers(
        daily_headers=list(DAILY_IMPORT_HEADERS),
        dream_headers=list(DREAM_IMPORT_HEADERS),
        daily_rows=daily_rows,
        dream_rows=dream_rows,
    )
    manifest = {
        'package_type': 'openmynd_export',
        'version': 1,
        'generated_at': '2026-06-10T00:00:00Z',
        'assets': manifest_assets or {},
    }
    manifest.update(manifest_overrides or {})

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, mode='w', compression=zipfile.ZIP_DEFLATED) as zf:
        zf.writestr('entries.xlsx', workbook_bytes)
        zf.writestr('manifest.json', json.dumps(manifest))
        for path, content in (media_files or {}).items():
            zf.writestr(path, content)
    return buf.getvalue()


def _make_all_data_zip_package(
    *,
    important_day_rows=None,
    thought_record_rows=None,
    manifest_assets=None,
    media_files=None,
) -> bytes:
    wb = openpyxl.Workbook()
    ws_daily = wb.active
    ws_daily.title = 'Daily'
    ws_daily.append(list(DAILY_IMPORT_HEADERS))
    ws_dreams = wb.create_sheet('Dreams')
    ws_dreams.append(list(DREAM_IMPORT_HEADERS))
    ws_important_days = wb.create_sheet('Important Days')
    ws_important_days.append(list(IMPORTANT_DAY_IMPORT_HEADERS))
    for row in important_day_rows or []:
        ws_important_days.append(row)
    ws_thought_records = wb.create_sheet('Thought Records')
    ws_thought_records.append(list(THOUGHT_RECORD_IMPORT_HEADERS))
    for row in thought_record_rows or []:
        ws_thought_records.append(row)

    workbook_buffer = io.BytesIO()
    wb.save(workbook_buffer)

    package_buffer = io.BytesIO()
    with zipfile.ZipFile(package_buffer, mode='w', compression=zipfile.ZIP_DEFLATED) as zf:
        zf.writestr('entries.xlsx', workbook_buffer.getvalue())
        zf.writestr(
            'manifest.json',
            json.dumps({
                'package_type': 'openmynd_export',
                'version': 1,
                'generated_at': '2026-08-06T00:00:00Z',
                'assets': manifest_assets or {},
            }),
        )
        for path, content in (media_files or {}).items():
            zf.writestr(path, content)
    return package_buffer.getvalue()


def _tiny_png_bytes() -> bytes:
    return base64.b64decode(
        'iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mP8/x8AAusB9Wn0nS8AAAAASUVORK5CYII='
    )


def _tiny_jpeg_bytes() -> bytes:
    return b'\xff\xd8\xff\xd9'


def _make_daylio_backup(*, include_photo: bool = False) -> bytes:
    photo_id = 301
    photo_checksum = 'synthetic-photo-checksum'
    backup = {
        'dayEntries': [{
            'note_title': 'Native backup entry',
            'note': 'Imported from a native backup.',
            'tags': [201],
            'assets': [photo_id] if include_photo else [],
            'hour': 20,
            'minute': 35,
            'day': 19,
            'month': 6,
            'year': 2026,
            'mood': 101,
        }],
        'customMoods': [{
            'id': 101,
            'custom_name': 'good',
        }],
        'tags': [{
            'id': 201,
            'name': 'friends',
        }],
        'assets': [{
            'id': photo_id,
            'checksum': photo_checksum,
            'type': 1,
        }] if include_photo else [],
    }
    payload = base64.b64encode(json.dumps(backup).encode('utf-8'))
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, mode='w', compression=zipfile.ZIP_DEFLATED) as zf:
        zf.writestr('backup.daylio', payload)
        if include_photo:
            zf.writestr(f'assets/photos/2026/7/{photo_checksum}', _tiny_jpeg_bytes())
    return buf.getvalue()


def _upload(client, token: str, file_bytes: bytes, filename='test.xlsx',
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            source='aidiary'):
    """POST multipart file to /api/import/upload."""
    return client.post(
        '/api/import/upload',
        headers={'Authorization': f'Bearer {token}'},
        data={
            'file': (io.BytesIO(file_bytes), filename, content_type),
            'source': source,
        },
        content_type='multipart/form-data',
    )


def _commit_review(client, token: str, import_session_id: str,
                   accepted_duplicate_row_ids=None, selected_row_ids=None,
                   entry_type_overrides=None):
    return client.post(
        '/api/import/commit',
        headers={'Authorization': f'Bearer {token}'},
        data=json.dumps({
            'import_session_id': import_session_id,
            'accepted_duplicate_row_ids': accepted_duplicate_row_ids or [],
            'selected_row_ids': selected_row_ids,
            'entry_type_overrides': entry_type_overrides or {},
        }),
        content_type='application/json',
    )


def _commit_all_review_entries(client, token: str, upload_response):
    data = json.loads(upload_response.data)
    review_entries = data.get('review_entries', [])
    selected = [row['row_id'] for row in review_entries]
    duplicates = [row['row_id'] for row in review_entries if row.get('is_duplicate')]
    return _commit_review(
        client,
        token,
        data['import_session_id'],
        accepted_duplicate_row_ids=duplicates,
        selected_row_ids=selected,
    )


def _create_legacy_import_history_table(conn: sqlite3.Connection) -> None:
    conn.executescript('''
        DROP TABLE IF EXISTS import_history;
        CREATE TABLE import_history (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            filename TEXT NOT NULL
        );
    ''')
    conn.commit()


# ---------------------------------------------------------------------------
# File type / size validation
# ---------------------------------------------------------------------------

class TestFileValidation:
    def test_missing_file_part(self, client):
        token = _register_and_login(client)
        resp = client.post(
            '/api/import/upload',
            headers={'Authorization': f'Bearer {token}'},
        )
        assert resp.status_code == 400
        data = json.loads(resp.data)
        assert data['status'] == 'error'
        assert data['errors']

    def test_no_filename(self, client):
        token = _register_and_login(client)
        resp = client.post(
            '/api/import/upload',
            headers={'Authorization': f'Bearer {token}'},
            data={'file': (io.BytesIO(b''), '', 'application/octet-stream')},
            content_type='multipart/form-data',
        )
        assert resp.status_code == 400
        data = json.loads(resp.data)
        assert data['status'] == 'error'

    def test_wrong_extension_rejected(self, client):
        token = _register_and_login(client)
        resp = _upload(client, token, b'dummy content', filename='diary.csv',
                       content_type='text/csv')
        assert resp.status_code == 422
        data = json.loads(resp.data)
        assert data['status'] == 'error'
        # Error message must mention the invalid type
        combined = ' '.join(data['errors'])
        assert 'csv' in combined.lower() or 'invalid file type' in combined.lower()

    def test_txt_extension_rejected(self, client):
        token = _register_and_login(client)
        resp = _upload(client, token, b'plain text', filename='diary.txt',
                       content_type='text/plain')
        assert resp.status_code == 422
        data = json.loads(resp.data)
        assert data['status'] == 'error'

    def test_oversized_file_rejected(self, client):
        from services.import_service import MAX_FILE_SIZE_BYTES
        token = _register_and_login(client)
        big_bytes = b'x' * (MAX_FILE_SIZE_BYTES + 1)
        resp = _upload(client, token, big_bytes, filename='big.xlsx')
        assert resp.status_code == 422
        data = json.loads(resp.data)
        assert data['status'] == 'error'
        combined = ' '.join(data['errors'])
        assert 'size' in combined.lower() or 'limit' in combined.lower()

    def test_empty_file_rejected(self, client):
        token = _register_and_login(client)
        resp = _upload(client, token, b'', filename='empty.xlsx')
        assert resp.status_code == 422
        data = json.loads(resp.data)
        assert data['status'] == 'error'

    def test_xlsx_extension_accepted(self, client):
        token = _register_and_login(client)
        file_bytes = _make_xlsx(daily_rows=[['2024-01-10', 'Day one', 'Good day', 'mood']])
        resp = _upload(client, token, file_bytes, filename='diary.xlsx')
        assert resp.status_code == 200
        data = json.loads(resp.data)
        assert data['status'] in ('success', 'empty', 'skipped')

    def test_unauthenticated_upload_rejected(self, client):
        file_bytes = _make_xlsx()
        resp = client.post(
            '/api/import/upload',
            data={'file': (io.BytesIO(file_bytes), 'test.xlsx')},
            content_type='multipart/form-data',
        )
        assert resp.status_code == 401


# ---------------------------------------------------------------------------
# Successful import
# ---------------------------------------------------------------------------

class TestSuccessfulImport:
    def test_nltk_rejects_mood_and_event_phrases_as_people_or_places(self):
        derived = derive_daily_nltk_fields(
            'Dinner invitation',
            'Good feelings about a dinner invitation.',
            excluded_terms={'daylio'},
        )

        assert derived['daily_people_names'] == ''
        assert derived['daily_places'] == ''

    def test_daylio_csv_import_maps_note_mood_activities_and_time(self, client):
        token = _register_and_login(client)
        csv_bytes = (
            'full_date,date,weekday,time,mood,activities,note_title,note\n'
            '2026-07-19,19 July 2026,Sunday,8:35 PM,good,walk|friends,Evening walk,Had a calm evening.\n'
        ).encode()

        resp = _upload(
            client,
            token,
            csv_bytes,
            filename='daylio_export.csv',
            content_type='text/csv',
            source='daylio',
        )
        resp = _commit_all_review_entries(client, token, resp)
        assert resp.status_code == 200
        data = json.loads(resp.data)
        assert data['summary']['inserted_daily'] == 1

        conn = sqlite3.connect(os.environ['DB_PATH'])
        conn.row_factory = sqlite3.Row
        row = conn.execute(
            '''SELECT entry_date, entry_time, title, user_message, mood, tags, import_id
               FROM dailydiary_entries WHERE title = ?''',
            ('Evening walk',),
        ).fetchone()
        conn.close()
        assert row['entry_date'] == '2026-07-19'
        assert row['entry_time'] == '20:35'
        assert row['user_message'] == 'Had a calm evening.'
        assert row['mood'] == 'good'
        assert {'walk', 'friends'}.issubset(set(row['tags'].split(',')))
        assert row['import_id'] == data['import_id']

    def test_daylio_title_and_mood_remain_separate(self, client):
        token = _register_and_login(client)
        csv_bytes = (
            'full_date,time,mood,activities,note_title,note\n'
            '2026-07-16,19:05,good,walking,Very good,A Daylio export entry.\n'
        ).encode()

        resp = _upload(
            client,
            token,
            csv_bytes,
            filename='daylio_export.csv',
            content_type='text/csv',
            source='daylio',
        )
        resp = _commit_all_review_entries(client, token, resp)

        assert resp.status_code == 200
        conn = sqlite3.connect(os.environ['DB_PATH'])
        row = conn.execute(
            '''SELECT title, mood, tags, daily_people_names, daily_places
               FROM dailydiary_entries WHERE entry_date = '2026-07-16' ''',
        ).fetchone()
        conn.close()
        assert row[0] == 'Very good'
        assert row[1] == 'good'
        assert all('daylio' not in (value or '').lower() for value in row[2:])

    def test_daylio_uses_mood_as_fallback_title_and_skips_checkins_without_notes(self, client):
        token = _register_and_login(client)
        csv_bytes = (
            'full_date,time,mood,activities,note_title,note\n'
            '2026-07-16,19:05,very good,walking,,An authored Daylio note.\n'
            '2026-07-16,21:10,quiet,reading,,\n'
        ).encode()

        preview = _upload(
            client,
            token,
            csv_bytes,
            filename='daylio_export.csv',
            content_type='text/csv',
            source='daylio',
        )

        assert preview.status_code == 200
        data = json.loads(preview.data)
        assert len(data['review_entries']) == 1
        assert data['review_entries'][0]['title'] == 'Very good'
        assert data['review_entries'][0]['mood'] == 'very good'
        assert any(
            'Skipped 1 Daylio mood/activity check-ins without authored notes.' in warning
            for warning in data['warnings']
        )

        committed = _commit_all_review_entries(client, token, preview)
        assert committed.status_code == 200
        conn = sqlite3.connect(os.environ['DB_PATH'])
        rows = conn.execute(
            'SELECT entry_time, title, user_message, mood FROM dailydiary_entries',
        ).fetchall()
        conn.close()
        assert rows == [('19:05', 'Very good', 'An authored Daylio note.', 'very good')]

    def test_daylio_same_day_matching_content_remains_distinct_at_different_times(self, client):
        token = _register_and_login(client)
        preview = _upload(
            client,
            token,
            (
                'full_date,time,mood,note_title,note\n'
                '2026-07-16,09:05,good,Repeated thought,Same authored text.\n'
                '2026-07-16,21:10,good,Repeated thought,Same authored text.\n'
            ).encode(),
            filename='daylio_export.csv',
            content_type='text/csv',
            source='daylio',
        )

        assert preview.status_code == 200
        data = json.loads(preview.data)
        assert len(data['review_entries']) == 2
        assert data['summary']['duplicate_daily'] == 0

    def test_daylio_html_notes_are_converted_to_plain_text(self, client):
        token = _register_and_login(client)
        preview = _upload(
            client,
            token,
            (
                'full_date,time,mood,note_title,note\n'
                '2026-07-16,19:05,good,Formatted note,'
                '"<p style=""font-style: normal; font-size: 26px"">Readable '
                '<strong>entry text</strong></p>"\n'
            ).encode(),
            filename='daylio_export.csv',
            content_type='text/csv',
            source='daylio',
        )

        assert preview.status_code == 200
        data = json.loads(preview.data)
        assert data['review_entries'][0]['content_preview'] == 'Readable entry text'

        committed = _commit_all_review_entries(client, token, preview)
        assert committed.status_code == 200
        conn = sqlite3.connect(os.environ['DB_PATH'])
        stored_text = conn.execute(
            "SELECT user_message FROM dailydiary_entries WHERE title = 'Formatted note'",
        ).fetchone()[0]
        conn.close()
        assert stored_text == 'Readable entry text'

    def test_reviewed_import_can_run_as_background_job(self, client):
        token = _register_and_login(client)
        preview = _upload(
            client,
            token,
            (
                'full_date,time,mood,note_title,note\n'
                '2026-07-14,09:00,good,Morning note,First entry.\n'
                '2026-07-14,21:00,quiet,Evening note,Second entry.\n'
            ).encode(),
            filename='daylio_export.csv',
            content_type='text/csv',
            source='daylio',
        )
        preview_data = json.loads(preview.data)
        selected_ids = [row['row_id'] for row in preview_data['review_entries']]

        started = client.post(
            '/api/import/jobs',
            headers={'Authorization': f'Bearer {token}'},
            data=json.dumps({
                'import_session_id': preview_data['import_session_id'],
                'accepted_duplicate_row_ids': [],
                'selected_row_ids': selected_ids,
                'entry_type_overrides': {},
            }),
            content_type='application/json',
        )
        assert started.status_code == 202
        job = json.loads(started.data)

        for _ in range(100):
            status_response = client.get(
                f'/api/import/jobs/{job["id"]}',
                headers={'Authorization': f'Bearer {token}'},
            )
            assert status_response.status_code == 200
            job = json.loads(status_response.data)
            if job['status'] in {'completed', 'failed'}:
                break
            time.sleep(0.02)

        assert job['status'] == 'completed'
        assert job['processed'] == 2
        assert job['total'] == 2
        assert job['percent'] == 100
        assert job['result']['summary']['inserted_daily'] == 2
        conn = sqlite3.connect(os.environ['DB_PATH'])
        persisted_job = conn.execute(
            '''SELECT status, processed, total, result_json, request_json, attempt_count
               FROM import_jobs WHERE id = ?''',
            (job['id'],),
        ).fetchone()
        conn.close()
        assert persisted_job[0:3] == ('completed', 2, 2)
        assert json.loads(persisted_job[3])['summary']['inserted_daily'] == 2
        assert json.loads(persisted_job[4])['selected_row_ids'] == selected_ids
        assert persisted_job[5] == 1

    def test_reviewed_import_completes_within_vercel_request(self, client, monkeypatch):
        monkeypatch.setenv('VERCEL', '1')
        token = _register_and_login(client)
        preview = _upload(
            client,
            token,
            (
                'full_date,time,mood,note_title,note\n'
                '2026-07-16,09:00,good,Vercel note,Stored before the response.\n'
            ).encode(),
            filename='daylio_export.csv',
            content_type='text/csv',
            source='daylio',
        )
        preview_data = json.loads(preview.data)
        selected_ids = [row['row_id'] for row in preview_data['review_entries']]

        started = client.post(
            '/api/import/jobs',
            headers={'Authorization': f'Bearer {token}'},
            data=json.dumps({
                'import_session_id': preview_data['import_session_id'],
                'accepted_duplicate_row_ids': [],
                'selected_row_ids': selected_ids,
                'entry_type_overrides': {},
            }),
            content_type='application/json',
        )

        assert started.status_code == 202
        job = json.loads(started.data)
        assert job['status'] == 'completed'
        assert job['processed'] == 1
        assert job['result']['summary']['inserted_daily'] == 1

    def test_polling_recovers_a_durable_import_job_after_worker_lease_expires(self, client):
        token = _register_and_login(client)
        preview = _upload(
            client,
            token,
            (
                'full_date,time,mood,note_title,note\n'
                '2026-07-15,18:30,good,Recovered note,Stored queue request.\n'
            ).encode(),
            filename='daylio_export.csv',
            content_type='text/csv',
            source='daylio',
        )
        preview_data = json.loads(preview.data)
        selected_ids = [row['row_id'] for row in preview_data['review_entries']]
        job_id = 'durable-recovery-job'
        now = '2026-07-20T12:00:00Z'
        conn = sqlite3.connect(os.environ['DB_PATH'])
        conn.execute(
            '''INSERT INTO import_jobs
               (id, user_id, import_session_id, status, processed, total, percent,
                message, request_json, created_at, updated_at, worker_token,
                lease_expires_at)
               VALUES (?, 1, ?, 'running', 0, 1, 0, ?, ?, ?, ?, ?, ?)''',
            (
                job_id,
                preview_data['import_session_id'],
                'Import interrupted…',
                json.dumps({
                    'accepted_duplicate_row_ids': [],
                    'selected_row_ids': selected_ids,
                    'entry_type_overrides': {},
                }),
                now,
                now,
                'stale-worker',
                '2026-07-20T11:59:00Z',
            ),
        )
        conn.commit()
        conn.close()

        for _ in range(100):
            status_response = client.get(
                f'/api/import/jobs/{job_id}',
                headers={'Authorization': f'Bearer {token}'},
            )
            assert status_response.status_code == 200
            job = json.loads(status_response.data)
            if job['status'] in {'completed', 'failed'}:
                break
            time.sleep(0.02)

        assert job['status'] == 'completed'
        assert job['result']['summary']['inserted_daily'] == 1

    def test_vercel_status_poll_marks_stale_job_failed_without_relaunching(self, client, monkeypatch):
        monkeypatch.setenv('VERCEL', '1')
        token = _register_and_login(client)
        job_id = 'stale-vercel-import-job'
        stale_time = '2026-07-20T12:00:00Z'
        conn = sqlite3.connect(os.environ['DB_PATH'])
        conn.execute(
            '''INSERT INTO import_jobs
               (id, user_id, import_session_id, status, processed, total, percent,
                message, request_json, created_at, updated_at, worker_token,
                lease_expires_at)
               VALUES (?, 1, ?, 'running', 0, 1173, 0, ?, ?, ?, ?, ?, ?)''',
            (
                job_id,
                'stale-session',
                'Import queued…',
                '{}',
                stale_time,
                stale_time,
                'abandoned-worker',
                '2026-07-20T12:30:00Z',
            ),
        )
        conn.commit()
        conn.close()

        response = client.get(
            f'/api/import/jobs/{job_id}',
            headers={'Authorization': f'Bearer {token}'},
        )

        assert response.status_code == 200
        job = json.loads(response.data)
        assert job['status'] == 'failed'
        assert job['error'] == 'The import did not finish within the serverless request window. Start it again.'

    def test_daylio_csv_rejects_missing_date_header(self, client):
        token = _register_and_login(client)
        resp = _upload(
            client,
            token,
            b'mood,note\ngood,No date\n',
            filename='daylio_export.csv',
            content_type='text/csv',
            source='daylio',
        )
        assert resp.status_code == 422
        assert 'date' in ' '.join(json.loads(resp.data)['errors']).lower()

    def test_daylio_native_backup_imports_entries_and_photos(self, client):
        token = _register_and_login(client)
        resp = _upload(
            client,
            token,
            _make_daylio_backup(include_photo=True),
            filename='backup.daylio',
            content_type='application/octet-stream',
            source='daylio',
        )
        resp = _commit_all_review_entries(client, token, resp)
        assert resp.status_code == 200
        data = json.loads(resp.data)
        assert data['summary']['inserted_daily'] == 1

        conn = sqlite3.connect(os.environ['DB_PATH'])
        conn.row_factory = sqlite3.Row
        row = conn.execute(
            '''SELECT id, entry_date, entry_time, mood, tags
               FROM dailydiary_entries WHERE title = ?''',
            ('Native backup entry',),
        ).fetchone()
        attachment = conn.execute(
            '''SELECT original_filename, mime_type, asset_role
               FROM entry_assets WHERE entry_type = 'daily' AND entry_id = ?''',
            (row['id'],),
        ).fetchone()
        conn.close()

        assert row['entry_date'] == '2026-07-19'
        assert row['entry_time'] == '20:35'
        assert row['mood'] == 'good'
        assert 'friends' in row['tags'].split(',')
        assert attachment['original_filename'] == 'daylio-photo-301.jpg'
        assert attachment['mime_type'] == 'image/jpeg'
        assert attachment['asset_role'] == 'attachment'

    def test_daylio_native_backup_rejects_invalid_archive(self, client):
        token = _register_and_login(client)
        resp = _upload(
            client,
            token,
            b'not a native backup',
            filename='backup.daylio',
            content_type='application/octet-stream',
            source='daylio',
        )
        assert resp.status_code == 422
        assert 'valid .daylio archive' in ' '.join(json.loads(resp.data)['errors'])

    def test_completed_import_can_be_reverted_with_its_attachments(self, client):
        token = _register_and_login(client)
        upload_preview = _upload(
            client,
            token,
            _make_daylio_backup(include_photo=True),
            filename='backup.daylio',
            content_type='application/octet-stream',
            source='daylio',
        )
        upload = _commit_all_review_entries(client, token, upload_preview)
        import_id = json.loads(upload.data)['import_id']

        response = client.post(
            f'/api/import/history/{import_id}/revert',
            headers={'Authorization': f'Bearer {token}'},
            data=json.dumps({'confirmation_text': 'REVERT IMPORT'}),
            content_type='application/json',
        )

        assert response.status_code == 200
        assert json.loads(response.data)['deleted_total'] == 1
        conn = sqlite3.connect(os.environ['DB_PATH'])
        assert conn.execute(
            'SELECT COUNT(*) FROM dailydiary_entries WHERE import_id = ?',
            (import_id,),
        ).fetchone()[0] == 0
        assert conn.execute(
            'SELECT status FROM import_history WHERE id = ?',
            (import_id,),
        ).fetchone()[0] == 'reverted'
        assert conn.execute('SELECT COUNT(*) FROM entry_assets').fetchone()[0] == 0
        conn.close()

    def test_selected_entry_deletion_is_explicit_and_user_scoped(self, client):
        token = _register_and_login(client)
        csv_bytes = (
            'full_date,time,mood,note_title,note\n'
            '2026-07-14,19:00,good,Keep this,First entry.\n'
            '2026-07-15,19:00,good,Delete this,Second entry.\n'
        ).encode()
        upload_preview = _upload(
            client,
            token,
            csv_bytes,
            filename='daylio.csv',
            content_type='text/csv',
            source='daylio',
        )
        assert upload_preview.status_code == 200
        assert _commit_all_review_entries(client, token, upload_preview).status_code == 200
        conn = sqlite3.connect(os.environ['DB_PATH'])
        delete_id = conn.execute(
            "SELECT id FROM dailydiary_entries WHERE title = 'Delete this'",
        ).fetchone()[0]
        conn.close()

        response = client.post(
            '/api/entries/delete-selected',
            headers={'Authorization': f'Bearer {token}'},
            data=json.dumps({'entries': [{'type': 'daily', 'id': delete_id}]}),
            content_type='application/json',
        )

        assert response.status_code == 200
        assert json.loads(response.data)['deleted_total'] == 1
        conn = sqlite3.connect(os.environ['DB_PATH'])
        titles = {
            row[0] for row in conn.execute('SELECT title FROM dailydiary_entries').fetchall()
        }
        conn.close()
        assert titles == {'Keep this'}

    def test_external_import_commits_only_selected_review_rows(self, client):
        token = _register_and_login(client)
        preview = _upload(
            client,
            token,
            (
                'full_date,time,mood,note_title,note\n'
                '2026-07-12,19:00,good,Include this,Selected entry.\n'
                '2026-07-13,19:00,good,Leave this out,Unselected entry.\n'
            ).encode(),
            filename='daylio.csv',
            content_type='text/csv',
            source='daylio',
        )
        preview_data = json.loads(preview.data)
        assert preview_data['status'] == 'review_required'
        assert len(preview_data['review_entries']) == 2
        selected_id = next(
            row['row_id'] for row in preview_data['review_entries']
            if row['title'] == 'Include this'
        )

        committed = _commit_review(
            client,
            token,
            preview_data['import_session_id'],
            selected_row_ids=[selected_id],
        )

        assert committed.status_code == 200
        conn = sqlite3.connect(os.environ['DB_PATH'])
        titles = {
            row[0] for row in conn.execute('SELECT title FROM dailydiary_entries').fetchall()
        }
        conn.close()
        assert titles == {'Include this'}

    def test_daylio_review_can_convert_daily_candidate_to_dream(self, client):
        token = _register_and_login(client)
        preview = _upload(
            client,
            token,
            (
                'full_date,time,mood,note_title,note\n'
                '2026-07-11,08:15,uneasy,Night train,I dreamed about a train.\n'
            ).encode(),
            filename='daylio.csv',
            content_type='text/csv',
            source='daylio',
        )
        preview_data = json.loads(preview.data)
        row_id = preview_data['review_entries'][0]['row_id']

        committed = _commit_review(
            client,
            token,
            preview_data['import_session_id'],
            selected_row_ids=[row_id],
            entry_type_overrides={row_id: 'dream'},
        )

        assert committed.status_code == 200
        conn = sqlite3.connect(os.environ['DB_PATH'])
        dream = conn.execute(
            'SELECT title, plot, emotion FROM dreamdiary_entries',
        ).fetchone()
        daily_count = conn.execute('SELECT COUNT(*) FROM dailydiary_entries').fetchone()[0]
        conn.close()
        assert dream == ('Night train', 'I dreamed about a train.', 'uneasy')
        assert daily_count == 0

    def test_daylio_duplicate_uses_staged_review(self, client):
        token = _register_and_login(client)
        csv_bytes = (
            'full_date,time,mood,activities,note_title,note\n'
            '2026-07-18,19:10,good,reading,Quiet night,Read a book.\n'
        ).encode()
        upload_args = {
            'filename': 'daylio_export.csv',
            'content_type': 'text/csv',
            'source': 'daylio',
        }

        first_preview = _upload(client, token, csv_bytes, **upload_args)
        first = _commit_all_review_entries(client, token, first_preview)
        second = _upload(client, token, csv_bytes, **upload_args)

        assert first.status_code == 200
        second_data = json.loads(second.data)
        assert second_data['status'] == 'review_required'
        assert second_data['summary']['duplicate_daily'] == 1
        assert second_data['import_session_id']

    def test_daylio_csv_accepts_semicolon_delimiter_and_combined_datetime(self, client):
        token = _register_and_login(client)
        csv_bytes = (
            'full_date;mood;activities;note_title;note\n'
            '2026-07-17T21:45;meh;rest|music;Quiet evening;Stayed home.\n'
        ).encode()

        resp = _upload(
            client,
            token,
            csv_bytes,
            filename='daylio_export.csv',
            content_type='text/csv',
            source='daylio',
        )
        resp = _commit_all_review_entries(client, token, resp)

        assert resp.status_code == 200
        conn = sqlite3.connect(os.environ['DB_PATH'])
        row = conn.execute(
            'SELECT entry_date, entry_time FROM dailydiary_entries WHERE title = ?',
            ('Quiet evening',),
        ).fetchone()
        conn.close()
        assert row == ('2026-07-17', '21:45')

    def test_daily_entries_inserted(self, client):
        token = _register_and_login(client)
        file_bytes = _make_xlsx(daily_rows=[
            ['2024-02-01', 'Morning', 'Woke up early', 'morning,routine'],
            ['2024-02-02', 'Afternoon', 'Had lunch', 'food'],
        ])
        resp = _upload(client, token, file_bytes)
        assert resp.status_code == 200
        data = json.loads(resp.data)
        assert data['status'] == 'success'
        assert data['summary']['inserted_daily'] == 2
        assert data['summary']['skipped_daily'] == 0
        assert data['import_id'] is not None

    def test_dream_entries_inserted(self, client):
        token = _register_and_login(client)
        file_bytes = _make_xlsx(dream_rows=[
            ['2024-03-01', 'Flying', 'I was flying', '', '', '', 'joy', '', '', '', '', ''],
        ])
        resp = _upload(client, token, file_bytes)
        assert resp.status_code == 200
        data = json.loads(resp.data)
        assert data['status'] == 'success'
        assert data['summary']['inserted_dreams'] == 1

    def test_mixed_entries_inserted(self, client):
        token = _register_and_login(client)
        file_bytes = _make_xlsx(
            daily_rows=[['2024-04-01', 'Entry 1', 'Content', 'tag1']],
            dream_rows=[['2024-04-02', 'Dream 1', 'Plot', '', '', '', '', '', '', '', '', '']],
        )
        resp = _upload(client, token, file_bytes)
        assert resp.status_code == 200
        data = json.loads(resp.data)
        assert data['summary']['inserted_daily'] == 1
        assert data['summary']['inserted_dreams'] == 1

    def test_response_shape(self, client):
        """Verify the API contract: all required fields are present."""
        token = _register_and_login(client)
        file_bytes = _make_xlsx(daily_rows=[['2024-05-01', 'Shape test', 'body', '']])
        resp = _upload(client, token, file_bytes)
        data = json.loads(resp.data)

        assert 'status' in data
        assert 'summary' in data
        assert 'warnings' in data
        assert 'import_id' in data
        assert 'import_session_id' in data

        summary = data['summary']
        for key in ('inserted_daily', 'skipped_daily',
                    'inserted_dreams', 'skipped_dreams',
                    'duplicate_dates_daily', 'duplicate_dates_dreams'):
            assert key in summary, f'Missing key in summary: {key}'

        assert isinstance(data['warnings'], list)
        assert 'duplicate_entries' in data
        assert isinstance(data['duplicate_entries'], list)


class TestImportIntegration:
    def test_imported_entries_expose_source_metadata_on_detail(self, client):
        token = _register_and_login(client)
        file_bytes = _make_xlsx(
            daily_rows=[['2025-03-01', 'Daily source', 'Daily body', 'Daily AI']],
            dream_rows=[
                ['2025-03-02', 'Dream source', 'Dream plot', '', '', '', '', '', '', '', '', '']
            ],
        )

        import_resp = _upload(
            client,
            token,
            file_bytes,
            filename='source-package.xlsx',
        )
        assert import_resp.status_code == 200
        import_id = json.loads(import_resp.data)['import_id']

        headers = {'Authorization': f'Bearer {token}'}
        daily_entries = json.loads(client.get('/api/daily', headers=headers).data)
        dream_entries = json.loads(client.get('/api/dreams', headers=headers).data)

        daily_detail = json.loads(
            client.get(f"/api/daily/{daily_entries[0]['id']}", headers=headers).data
        )
        dream_detail = json.loads(
            client.get(f"/api/dreams/{dream_entries[0]['id']}", headers=headers).data
        )

        for detail in (daily_detail, dream_detail):
            assert detail['import_id'] == import_id
            assert detail['import_metadata']['filename'] == 'source-package.xlsx'
            assert detail['import_metadata']['imported_at']

    def test_imported_daily_entries_readable(self, client):
        token = _register_and_login(client)
        target_date = '2025-04-01'
        target_title = 'Imported Daily Smoke'
        file_bytes = _make_xlsx(
            daily_rows=[[target_date, target_title, 'Imported body text', 'Imported AI text']]
        )

        import_resp = _upload(client, token, file_bytes)
        assert import_resp.status_code == 200

        list_resp = client.get(
            '/api/daily',
            headers={'Authorization': f'Bearer {token}'},
        )
        assert list_resp.status_code == 200
        entries = json.loads(list_resp.data)
        assert any(
            entry.get('entry_date') == target_date or entry.get('title') == target_title
            for entry in entries
        )

    def test_imported_dreams_entries_readable(self, client):
        token = _register_and_login(client)
        target_date = '2025-04-02'
        target_title = 'Imported Dream Smoke'
        file_bytes = _make_xlsx(
            dream_rows=[[
                target_date,
                target_title,
                'I was exploring a forest',
                'Guide',
                'Forest',
                'Unknown',
                'curious',
                'trees',
                'stay grounded',
                'journal',
                'none',
                'smoke',
            ]]
        )

        import_resp = _upload(client, token, file_bytes)
        assert import_resp.status_code == 200

        list_resp = client.get(
            '/api/dreams',
            headers={'Authorization': f'Bearer {token}'},
        )
        assert list_resp.status_code == 200
        entries = json.loads(list_resp.data)
        assert any(
            entry.get('entry_date') == target_date or entry.get('title') == target_title
            for entry in entries
        )


# ---------------------------------------------------------------------------
# Duplicate handling
# ---------------------------------------------------------------------------

class TestDuplicateHandling:
    def test_duplicate_daily_entry_same_day_same_title_and_content_requires_review(self, client):
        token = _register_and_login(client)
        file_bytes = _make_xlsx(daily_rows=[['2024-06-01', 'First', 'Body', '']])
        # First upload
        _upload(client, token, file_bytes)
        # Second upload — same date, same title, same content
        resp = _upload(client, token, file_bytes)
        data = json.loads(resp.data)
        assert data['status'] == 'review_required'
        assert data['summary']['duplicate_daily'] == 1
        assert data['import_session_id']

        conn = sqlite3.connect(os.environ['DB_PATH'])
        count = conn.execute(
            "SELECT COUNT(*) FROM dailydiary_entries WHERE entry_date = '2024-06-01'"
        ).fetchone()[0]
        conn.close()
        assert count == 1

    def test_same_day_different_title_is_allowed(self, client):
        token = _register_and_login(client)
        first_file = _make_xlsx(daily_rows=[['2024-06-01', 'Morning notes', 'Body', '']])
        second_file = _make_xlsx(daily_rows=[['2024-06-01', 'Evening notes', 'Body', '']])

        first_resp = _upload(client, token, first_file)
        second_resp = _upload(client, token, second_file)

        assert first_resp.status_code == 200
        data = json.loads(second_resp.data)
        assert data['summary']['inserted_daily'] == 1
        assert data['summary']['skipped_daily'] == 0

    def test_same_day_same_title_different_content_is_allowed(self, client):
        token = _register_and_login(client)
        first_file = _make_xlsx(daily_rows=[['2024-06-01', 'Gym twice', 'Morning gym session', '']])
        second_file = _make_xlsx(daily_rows=[['2024-06-01', 'Gym twice', 'Evening gym session', '']])

        _upload(client, token, first_file)
        second_resp = _upload(client, token, second_file)

        data = json.loads(second_resp.data)
        assert data['summary']['inserted_daily'] == 1
        assert data['summary']['skipped_daily'] == 0

    def test_duplicate_dream_same_day_same_title_and_plot_requires_review(self, client):
        token = _register_and_login(client)
        dream = [
            '2024-06-02', 'Flying again', 'I flew over the same city', '', '', '',
            'joy', '', '', '', '', '',
        ]
        file_bytes = _make_xlsx(dream_rows=[dream])

        _upload(client, token, file_bytes)
        response = _upload(client, token, file_bytes)
        data = json.loads(response.data)

        assert data['status'] == 'review_required'
        assert data['summary']['duplicate_dreams'] == 1
        assert data['duplicate_entries'][0]['entry_type'] == 'dream'
        assert data['duplicate_entries'][0]['reason'] == 'same_date_time_title_content'

    def test_same_day_dreams_with_different_plots_are_allowed(self, client):
        token = _register_and_login(client)
        first = [
            '2024-06-03', 'Recurring place', 'I entered a quiet library', '', '', '',
            'calm', '', '', '', '', '',
        ]
        second = [
            '2024-06-03', 'Recurring place', 'I ran through a crowded station', '', '', '',
            'urgent', '', '', '', '', '',
        ]

        _upload(client, token, _make_xlsx(dream_rows=[first]))
        response = _upload(client, token, _make_xlsx(dream_rows=[second]))
        data = json.loads(response.data)

        assert data['summary']['inserted_dreams'] == 1
        assert data['summary']['skipped_dreams'] == 0

    def test_duplicate_reported_in_warnings(self, client):
        token = _register_and_login(client)
        file_bytes = _make_xlsx(daily_rows=[['2024-07-01', 'T', 'B', '']])
        _upload(client, token, file_bytes)
        resp = _upload(client, token, file_bytes)
        data = json.loads(resp.data)
        assert data['status'] == 'review_required'
        assert data['message'] == 'Duplicates found. Review and confirm before importing.'

    def test_duplicate_entries_payload_includes_date_title_type_and_reason(self, client):
        token = _register_and_login(client)
        file_bytes = _make_xlsx(daily_rows=[['2024-07-02', 'Gym twice', 'Morning session', '']])
        _upload(client, token, file_bytes)
        resp = _upload(client, token, file_bytes)
        data = json.loads(resp.data)

        assert data['duplicate_entries']
        duplicate = data['duplicate_entries'][0]
        assert duplicate['row_id'] == 'daily-1'
        assert duplicate['entry_type'] == 'daily'
        assert duplicate['entry_date'] == '2024-07-02'
        assert duplicate['title'] == 'Gym twice'
        assert duplicate['reason'] == 'same_date_time_title_content'
        assert 'Morning session' in duplicate['content_preview']

    def test_partial_duplicate(self, client):
        """One existing same-day/same-title/same-content row + one new row → only new one inserted."""
        token = _register_and_login(client)
        file1 = _make_xlsx(daily_rows=[['2024-08-01', 'Existing', 'Body', '']])
        _upload(client, token, file1)

        file2 = _make_xlsx(daily_rows=[
            ['2024-08-01', 'Existing', 'Body', ''],
            ['2024-08-02', 'New', 'Should insert', ''],
        ])
        resp = _upload(client, token, file2)
        data = json.loads(resp.data)
        assert data['status'] == 'review_required'
        assert data['summary']['ready_daily'] == 1
        assert data['summary']['duplicate_daily'] == 1

        commit_resp = _commit_review(client, token, data['import_session_id'], [])
        commit_data = json.loads(commit_resp.data)
        assert commit_data['summary']['inserted_daily'] == 1
        assert commit_data['summary']['skipped_daily'] == 1

    def test_all_duplicates_returns_skipped_status(self, client):
        token = _register_and_login(client)
        file_bytes = _make_xlsx(daily_rows=[['2024-09-01', 'T', 'B', '']])
        _upload(client, token, file_bytes)
        resp = _upload(client, token, file_bytes)
        data = json.loads(resp.data)
        commit_resp = _commit_review(client, token, data['import_session_id'], [])
        commit_data = json.loads(commit_resp.data)
        assert commit_data['status'] == 'skipped'

    def test_commit_selected_duplicates_imports_with_duplicate_tag(self, client):
        token = _register_and_login(client)
        file_bytes = _make_xlsx(daily_rows=[['2024-09-03', 'Repeat', 'Body', '']])
        _upload(client, token, file_bytes)
        preview_resp = _upload(client, token, file_bytes)
        preview_data = json.loads(preview_resp.data)
        duplicate_row = preview_data['duplicate_entries'][0]

        commit_resp = _commit_review(
            client,
            token,
            preview_data['import_session_id'],
            [duplicate_row['row_id']],
        )
        commit_data = json.loads(commit_resp.data)
        assert commit_data['status'] == 'success'

        conn = sqlite3.connect(os.environ['DB_PATH'])
        conn.row_factory = sqlite3.Row
        rows = conn.execute(
            "SELECT tags, import_id FROM dailydiary_entries WHERE entry_date = '2024-09-03' ORDER BY id DESC"
        ).fetchall()
        conn.close()
        assert len(rows) == 2
        assert '*Duplicate*' in (rows[0]['tags'] or '')
        assert rows[0]['import_id'] == commit_data['import_id']


# ---------------------------------------------------------------------------
# Invalid / malformed data rows
# ---------------------------------------------------------------------------

class TestMalformedData:
    def test_invalid_date_row_skipped_with_warning(self, client):
        token = _register_and_login(client)
        file_bytes = _make_xlsx(daily_rows=[
            ['not-a-date', 'Bad row', 'Should be skipped', ''],
            ['2024-10-01', 'Good row', 'Should be inserted', ''],
        ])
        resp = _upload(client, token, file_bytes)
        data = json.loads(resp.data)
        assert data['summary']['inserted_daily'] == 1
        combined = ' '.join(data['warnings'])
        assert 'skipped' in combined.lower() or 'invalid' in combined.lower()

    def test_empty_workbook_no_data(self, client):
        """Workbook with correct sheets but no data rows → status empty."""
        token = _register_and_login(client)
        file_bytes = _make_xlsx()  # headers only, no data rows
        resp = _upload(client, token, file_bytes)
        assert resp.status_code == 200
        data = json.loads(resp.data)
        assert data['status'] == 'empty'
        assert data['summary']['inserted_daily'] == 0
        assert data['summary']['inserted_dreams'] == 0

    def test_injection_stripped_from_content(self, client):
        token = _register_and_login(client)
        file_bytes = _make_xlsx(daily_rows=[
            ['2024-11-01', '<script>alert(1)</script>', 'javascript:alert(1)', 'tag'],
        ])
        resp = _upload(client, token, file_bytes)
        assert resp.status_code == 200
        data = json.loads(resp.data)
        assert data['summary']['inserted_daily'] == 1
        # Verify database stored sanitised values
        db_path = os.environ['DB_PATH']
        conn = sqlite3.connect(db_path)
        conn.row_factory = sqlite3.Row
        row = conn.execute(
            "SELECT title, user_message FROM dailydiary_entries WHERE entry_date = '2024-11-01'"
        ).fetchone()
        conn.close()
        assert row is not None
        assert '<script>' not in row['title']
        assert 'javascript:' not in row['user_message']

    def test_import_preserves_entry_time_from_new_format_workbook(self, client):
        token = _register_and_login(client)
        file_bytes = _make_xlsx_with_headers(
            daily_headers=list(DAILY_IMPORT_HEADERS),
            dream_headers=list(DREAM_IMPORT_HEADERS),
            daily_rows=[['2024-11-02', '14:25', 'Timed daily', 'Body', 'AI']],
            dream_rows=[[
                '2024-11-03',
                '06:40',
                'Timed dream',
                'A long corridor',
                'Guide',
                'House',
                'Past',
                'Uneasy',
                'Mirrors',
                'Notice patterns',
                'Pause',
                'None',
                'dream',
            ]],
        )

        resp = _upload(client, token, file_bytes)
        assert resp.status_code == 200

        conn = sqlite3.connect(os.environ['DB_PATH'])
        conn.row_factory = sqlite3.Row
        daily_row = conn.execute(
            "SELECT entry_time FROM dailydiary_entries WHERE title = 'Timed daily'"
        ).fetchone()
        dream_row = conn.execute(
            "SELECT entry_time FROM dreamdiary_entries WHERE title = 'Timed dream'"
        ).fetchone()
        conn.close()

        assert daily_row['entry_time'] == '14:25'
        assert dream_row['entry_time'] == '06:40'

    def test_legacy_workbook_without_entry_time_defaults_by_entry_type(self, client):
        token = _register_and_login(client)
        file_bytes = _make_xlsx_with_headers(
            daily_headers=['date', 'title', 'user_entry', 'ai_response'],
            dream_headers=[
                'date',
                'title',
                'plot',
                'cast',
                'location',
                'period',
                'emotion',
                'symbols_and_imagery',
                'insight',
                'action',
                'other',
                'tags',
            ],
            daily_rows=[['2024-11-04', 'Legacy daily', 'Body', 'AI']],
            dream_rows=[[
                '2024-11-05',
                'Legacy dream',
                'Ocean tide',
                'Friend',
                'Beach',
                'Future',
                'Calm',
                'Moon',
                'Slow down',
                'Rest',
                'None',
                'dream',
            ]],
        )

        resp = _upload(client, token, file_bytes)
        assert resp.status_code == 200

        conn = sqlite3.connect(os.environ['DB_PATH'])
        conn.row_factory = sqlite3.Row
        daily_row = conn.execute(
            "SELECT entry_time FROM dailydiary_entries WHERE title = 'Legacy daily'"
        ).fetchone()
        dream_row = conn.execute(
            "SELECT entry_time FROM dreamdiary_entries WHERE title = 'Legacy dream'"
        ).fetchone()
        conn.close()

        assert daily_row['entry_time'] == '19:00'
        assert dream_row['entry_time'] == '08:00'

    def test_zip_package_import_restores_bundled_image_metadata(self, client):
        token = _register_and_login(client)
        asset_ref = 'daily_20241106_1'
        package_bytes = _make_zip_package(
            daily_rows=[['2024-11-06', '18:10', 'Packaged daily', 'Body text', 'AI text', asset_ref]],
            manifest_assets={
                asset_ref: {
                    'entry_type': 'daily',
                    'image_filename': 'hero.png',
                    'image_source': 'ai',
                    'image_position_x': '33',
                    'image_position_y': '77',
                    'image_prompt': 'City lights reflected on wet pavement',
                    'recycled_image_prompt': '',
                }
            },
            media_files={
                f'media/{asset_ref}/hero.png': _tiny_png_bytes(),
            },
        )

        resp = _upload(client, token, package_bytes, filename='package.zip', content_type='application/zip')
        assert resp.status_code == 200
        data = json.loads(resp.data)
        assert data['summary']['inserted_daily'] == 1

        conn = sqlite3.connect(os.environ['DB_PATH'])
        conn.row_factory = sqlite3.Row
        row = conn.execute(
            """
            SELECT entry_time, image_storage_key, image_source, image_position_x, image_position_y, image_prompt
            FROM dailydiary_entries
            WHERE title = 'Packaged daily'
            """
        ).fetchone()
        conn.close()

        assert row['entry_time'] == '18:10'
        assert row['image_storage_key']
        assert row['image_source'] == 'ai'
        assert row['image_position_x'] == '33'
        assert row['image_position_y'] == '77'
        assert row['image_prompt'] == 'City lights reflected on wet pavement'

    def test_zip_package_import_restores_attachment_assets(self, client):
        token = _register_and_login(client)
        asset_ref = 'dream_20241107_1'
        package_bytes = _make_zip_package(
            dream_rows=[[
                '2024-11-07', '07:40', 'Packaged dream', 'River crossing', 'Mara',
                'Bridge', 'Present', 'Uneasy', 'Mist', 'Face the change',
                'Kept walking', 'None', 'river', asset_ref,
            ]],
            manifest_assets={
                asset_ref: {
                    'entry_type': 'dream',
                    'image_filename': '',
                    'attachments': [
                        {
                            'package_filename': '01_note.pdf',
                            'original_filename': 'note.pdf',
                            'mime_type': 'application/pdf',
                            'asset_role': 'attachment',
                            'file_size_bytes': 14,
                            'sort_order': 0,
                        },
                        {
                            'package_filename': '02_audio.m4a',
                            'original_filename': 'audio.m4a',
                            'mime_type': 'audio/mp4',
                            'asset_role': 'attachment',
                            'file_size_bytes': 12,
                            'sort_order': 1,
                        },
                    ],
                }
            },
            media_files={
                f'media/{asset_ref}/01_note.pdf': b'%PDF-1.4 test\n',
                f'media/{asset_ref}/02_audio.m4a': b'FAKEAUDIO123',
            },
        )

        resp = _upload(client, token, package_bytes, filename='package.zip', content_type='application/zip')
        assert resp.status_code == 200
        data = json.loads(resp.data)
        assert data['summary']['inserted_dreams'] == 1

        conn = sqlite3.connect(os.environ['DB_PATH'])
        conn.row_factory = sqlite3.Row
        asset_rows = conn.execute(
            """
            SELECT original_filename, mime_type, asset_role, file_size_bytes, sort_order, storage_key
            FROM entry_assets
            WHERE entry_type = 'dream'
            ORDER BY sort_order ASC, id ASC
            """
        ).fetchall()
        conn.close()

        assert len(asset_rows) == 2
        assert asset_rows[0]['original_filename'] == 'note.pdf'
        assert asset_rows[0]['mime_type'] == 'application/pdf'
        assert asset_rows[0]['asset_role'] == 'attachment'
        assert asset_rows[0]['sort_order'] == 0
        assert asset_rows[0]['storage_key']
        assert asset_rows[1]['original_filename'] == 'audio.m4a'
        assert asset_rows[1]['mime_type'] == 'audio/mp4'
        assert asset_rows[1]['sort_order'] == 1


# ---------------------------------------------------------------------------
# Schema contract regression coverage (Issue #39)
# ---------------------------------------------------------------------------

class TestSchemaContractWarnings:
    def test_package_version_mismatch_is_reported_without_blocking_supported_fields(self, client):
        token = _register_and_login(client)
        package_bytes = _make_zip_package(
            daily_rows=[['2025-02-01', '19:00', 'Future package', 'Body', '', '']],
            manifest_overrides={'version': 99},
        )

        response = _upload(
            client,
            token,
            package_bytes,
            filename='future-package.zip',
            content_type='application/zip',
        )

        assert response.status_code == 200
        data = json.loads(response.data)
        assert data['summary']['inserted_daily'] == 1
        assert any('version 99' in warning.lower() for warning in data['warnings'])

    def test_daily_unexpected_column_emits_warning(self, client):
        token = _register_and_login(client)
        file_bytes = _make_xlsx_with_headers(
            daily_headers=['date', 'title', 'user_entry', 'ai_response', 'mood_score'],
            dream_headers=list(DREAM_IMPORT_HEADERS),
            daily_rows=[['2025-03-01', 'T', 'Body', 'Imported answer', '9']],
        )

        resp = _upload(client, token, file_bytes)
        assert resp.status_code == 422
        data = json.loads(resp.data)
        combined = ' '.join(data['errors']).lower()

        assert 'daily sheet' in combined
        assert 'unexpected columns' in combined
        assert 'mood_score' in combined

    def test_daily_missing_required_column_warning(self, client):
        token = _register_and_login(client)
        file_bytes = _make_xlsx_with_headers(
            daily_headers=['date', 'title', 'ai_response'],
            dream_headers=list(DREAM_IMPORT_HEADERS),
            daily_rows=[['2025-03-02', 'No user_entry column', 'Imported answer']],
        )

        resp = _upload(client, token, file_bytes)
        assert resp.status_code == 422
        data = json.loads(resp.data)
        combined = ' '.join(data['errors']).lower()

        assert 'daily sheet' in combined
        assert 'missing columns' in combined
        assert 'user_entry' in combined

    def test_daily_missing_date_leads_to_skipped_rows_zero_inserted(self, client):
        token = _register_and_login(client)
        file_bytes = _make_xlsx_with_headers(
            daily_headers=list(DAILY_IMPORT_HEADERS),
            dream_headers=list(DREAM_IMPORT_HEADERS),
            daily_rows=[[None, 'Missing date', 'Body', 'false']],
        )

        resp = _upload(client, token, file_bytes)
        assert resp.status_code == 200
        data = json.loads(resp.data)
        combined = ' '.join(data['warnings']).lower()

        assert data['summary']['inserted_daily'] == 0
        assert data['summary']['skipped_daily'] == 0
        assert 'daily sheet row' in combined
        assert 'invalid or missing date' in combined

    def test_dreams_unexpected_column_emits_warning(self, client):
        token = _register_and_login(client)
        file_bytes = _make_xlsx_with_headers(
            daily_headers=list(DAILY_IMPORT_HEADERS),
            dream_headers=[*DREAM_IMPORT_HEADERS, 'lucidity_level'],
            dream_rows=[[
                '2025-03-03', 'Dream', 'Plot', '', '', '', 'joy', '', '', '', '', '', 'high'
            ]],
        )

        resp = _upload(client, token, file_bytes)
        assert resp.status_code == 200
        data = json.loads(resp.data)
        combined = ' '.join(data['warnings']).lower()

        assert 'dreams sheet' in combined
        assert 'unexpected columns' in combined
        assert 'lucidity_level' in combined

    def test_dreams_missing_plot_warning(self, client):
        token = _register_and_login(client)
        file_bytes = _make_xlsx_with_headers(
            daily_headers=list(DAILY_IMPORT_HEADERS),
            dream_headers=[
                'date',
                'title',
                'cast',
                'location',
                'period',
                'emotion',
                'symbols_and_imagery',
                'insight',
                'action',
                'other',
                'tags',
            ],
            dream_rows=[['2025-03-04', 'Dream title', '', '', '', '', '', '', '', '', 'tag']],
        )

        resp = _upload(client, token, file_bytes)
        assert resp.status_code == 200
        data = json.loads(resp.data)
        combined = ' '.join(data['warnings']).lower()

        assert 'dreams sheet' in combined
        assert 'missing columns' in combined
        assert 'plot' in combined

    def test_warnings_payload_is_list_of_strings(self, client):
        token = _register_and_login(client)
        file_bytes = _make_xlsx_with_headers(
            daily_headers=list(DAILY_IMPORT_HEADERS),
            dream_headers=[*DREAM_IMPORT_HEADERS, 'bonus_column'],
            daily_rows=[['2025-03-05', 'Payload shape', 'Body', 'false']],
            dream_rows=[[
                '2025-03-05', 'Dream', 'Plot', '', '', '', '', '', '', '', '', '', 'extra'
            ]],
        )

        resp = _upload(client, token, file_bytes)
        assert resp.status_code == 200
        data = json.loads(resp.data)

        assert isinstance(data['warnings'], list)
        assert data['warnings'], 'Expected at least one warning for schema mismatch.'
        assert all(isinstance(item, str) for item in data['warnings'])

    def test_valid_workbook_with_new_daily_headers_imports_successfully(self, client):
        token = _register_and_login(client)
        file_bytes = _make_xlsx_with_headers(
            daily_headers=list(DAILY_IMPORT_HEADERS),
            dream_headers=list(DREAM_IMPORT_HEADERS),
            daily_rows=[['2025-03-06', 'Valid row', 'Body text', 'false']],
        )

        resp = _upload(client, token, file_bytes)
        assert resp.status_code == 200
        data = json.loads(resp.data)
        assert data['status'] == 'success'
        assert data['summary']['inserted_daily'] == 1

    def test_ai_response_blank_defaults_empty(self, client):
        token = _register_and_login(client)
        file_bytes = _make_xlsx_with_headers(
            daily_headers=list(DAILY_IMPORT_HEADERS),
            dream_headers=list(DREAM_IMPORT_HEADERS),
            daily_rows=[['2025-03-07', 'No AI', 'Body text', '']],
        )

        resp = _upload(client, token, file_bytes)
        assert resp.status_code == 200
        data = json.loads(resp.data)
        assert data['summary']['inserted_daily'] == 1

        db_path = os.environ['DB_PATH']
        conn = sqlite3.connect(db_path)
        conn.row_factory = sqlite3.Row
        row = conn.execute(
            "SELECT ai_response FROM dailydiary_entries WHERE entry_date = '2025-03-07'"
        ).fetchone()
        conn.close()

        assert row is not None
        assert (row['ai_response'] or '') == ''

    def test_ai_response_is_imported_directly(self, client):
        token = _register_and_login(client)
        file_bytes = _make_xlsx_with_headers(
            daily_headers=list(DAILY_IMPORT_HEADERS),
            dream_headers=list(DREAM_IMPORT_HEADERS),
            daily_rows=[["2025-03-08", 'Imported AI', 'Body text', 'From spreadsheet']],
        )

        resp = _upload(client, token, file_bytes)
        assert resp.status_code == 200
        data = json.loads(resp.data)
        assert data['summary']['inserted_daily'] == 1

        db_path = os.environ['DB_PATH']
        conn = sqlite3.connect(db_path)
        conn.row_factory = sqlite3.Row
        row = conn.execute(
            "SELECT ai_response FROM dailydiary_entries WHERE entry_date = '2025-03-08'"
        ).fetchone()
        conn.close()

        assert row is not None
        assert row['ai_response'] == 'From spreadsheet'


# ---------------------------------------------------------------------------
# Import history
# ---------------------------------------------------------------------------

class TestImportHistory:
    def test_history_empty_initially(self, client):
        token = _register_and_login(client)
        resp = client.get(
            '/api/import/history',
            headers={'Authorization': f'Bearer {token}'},
        )
        assert resp.status_code == 200
        data = json.loads(resp.data)
        assert 'history' in data
        assert data['history'] == []

    def test_history_recorded_after_upload(self, client):
        token = _register_and_login(client)
        file_bytes = _make_xlsx(daily_rows=[['2024-12-01', 'Entry', 'Body', '']])
        _upload(client, token, file_bytes, filename='myfile.xlsx')
        resp = client.get(
            '/api/import/history',
            headers={'Authorization': f'Bearer {token}'},
        )
        data = json.loads(resp.data)
        assert len(data['history']) == 1
        record = data['history'][0]
        assert record['filename'] == 'myfile.xlsx'
        assert record['inserted_daily'] == 1
        assert record['status'] == 'success'

    def test_history_row_shape(self, client):
        """Verify every expected field is present in history records."""
        token = _register_and_login(client)
        file_bytes = _make_xlsx(daily_rows=[['2024-12-02', 'T', 'B', '']])
        _upload(client, token, file_bytes)
        resp = client.get(
            '/api/import/history',
            headers={'Authorization': f'Bearer {token}'},
        )
        data = json.loads(resp.data)
        record = data['history'][0]
        required_keys = {
            'id', 'imported_at', 'filename', 'file_size_bytes',
            'inserted_daily', 'skipped_daily',
            'inserted_dreams', 'skipped_dreams',
            'warnings', 'status',
        }
        assert required_keys.issubset(record.keys()), (
            f"Missing keys: {required_keys - record.keys()}"
        )
        assert isinstance(record['warnings'], list)

    def test_history_multiple_uploads(self, client):
        token = _register_and_login(client)
        for day in ('2025-01-01', '2025-01-02', '2025-01-03'):
            file_bytes = _make_xlsx(daily_rows=[[day, 'T', 'B', '']])
            _upload(client, token, file_bytes)
        resp = client.get(
            '/api/import/history',
            headers={'Authorization': f'Bearer {token}'},
        )
        data = json.loads(resp.data)
        assert len(data['history']) == 3

    def test_history_unauthenticated(self, client):
        resp = client.get('/api/import/history')
        assert resp.status_code == 401

    def test_history_isolated_per_user(self, client):
        """Two users should each see only their own history."""
        # User 1
        client.post('/api/register',
                    data=json.dumps({'username': 'userA', 'password': 'pw12345a'}),
                    content_type='application/json')
        r = client.post('/api/login',
                        data=json.dumps({'username': 'userA', 'password': 'pw12345a'}),
                        content_type='application/json')
        token_a = json.loads(r.data)['token']

        # User 2
        client.post('/api/register',
                    data=json.dumps({'username': 'userB', 'password': 'pw12345a'}),
                    content_type='application/json')
        r = client.post('/api/login',
                        data=json.dumps({'username': 'userB', 'password': 'pw12345a'}),
                        content_type='application/json')
        token_b = json.loads(r.data)['token']

        # User A uploads
        file_bytes = _make_xlsx(daily_rows=[['2025-02-01', 'T', 'B', '']])
        _upload(client, token_a, file_bytes)

        # User B should see empty history
        resp_b = client.get('/api/import/history',
                            headers={'Authorization': f'Bearer {token_b}'})
        data_b = json.loads(resp_b.data)
        assert data_b['history'] == []

    def test_legacy_history_table_is_repaired_on_upload(self, client):
        token = _register_and_login(client)
        db_path = os.environ['DB_PATH']
        conn = sqlite3.connect(db_path)
        _create_legacy_import_history_table(conn)
        conn.close()

        file_bytes = _make_xlsx(daily_rows=[['2025-02-02', 'Legacy', 'Body', '']])
        resp = _upload(client, token, file_bytes)
        assert resp.status_code == 200
        data = json.loads(resp.data)
        assert data['status'] == 'success'

        history_resp = client.get(
            '/api/import/history',
            headers={'Authorization': f'Bearer {token}'},
        )
        history_data = json.loads(history_resp.data)
        assert len(history_data['history']) == 1
        assert history_data['history'][0]['imported_at']


# ---------------------------------------------------------------------------
# Template download
# ---------------------------------------------------------------------------

class TestTemplateDownload:
    def test_template_returns_xlsx(self, client):
        token = _register_and_login(client)
        resp = client.get(
            '/api/import/template',
            headers={'Authorization': f'Bearer {token}'},
        )
        assert resp.status_code == 200
        assert 'spreadsheet' in resp.content_type or 'octet-stream' in resp.content_type

    def test_template_has_correct_sheets(self, client):
        token = _register_and_login(client)
        resp = client.get(
            '/api/import/template',
            headers={'Authorization': f'Bearer {token}'},
        )
        wb = openpyxl.load_workbook(io.BytesIO(resp.data))
        assert 'Daily' in wb.sheetnames
        assert 'Dreams' in wb.sheetnames

    def test_template_has_daily_headers(self, client):
        token = _register_and_login(client)
        resp = client.get(
            '/api/import/template',
            headers={'Authorization': f'Bearer {token}'},
        )
        wb = openpyxl.load_workbook(io.BytesIO(resp.data))
        ws = wb['Daily']
        headers = [cell.value for cell in ws[1]]
        assert headers == list(DAILY_IMPORT_HEADERS)

    def test_template_has_dreams_headers(self, client):
        token = _register_and_login(client)
        resp = client.get(
            '/api/import/template',
            headers={'Authorization': f'Bearer {token}'},
        )
        wb = openpyxl.load_workbook(io.BytesIO(resp.data))
        ws = wb['Dreams']
        headers = [cell.value for cell in ws[1]]
        assert headers == list(DREAM_IMPORT_HEADERS)

    def test_template_unauthenticated(self, client):
        resp = client.get('/api/import/template')
        assert resp.status_code == 401


# ---------------------------------------------------------------------------
# Export download
# ---------------------------------------------------------------------------

class TestExportDownload:
    @staticmethod
    def _seed_export_rows(db_path: str, username: str = 'importer') -> None:
        conn = sqlite3.connect(db_path)
        conn.row_factory = sqlite3.Row

        user_id = conn.execute(
            "SELECT id FROM users WHERE username = ?", (username,)
        ).fetchone()['id']

        conn.execute(
            """
            INSERT INTO dailydiary_entries
            (user_id, entry_date, entry_time, entry_number, title, user_message, ai_response, tags)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (user_id, '2026-01-10', '09:30', 1, 'Daily one', 'Body text', 'AI text', 'tag1,tag2'),
        )
        conn.execute(
            """
            INSERT INTO dailydiary_entries
            (user_id, entry_date, entry_time, entry_number, title, user_message, ai_response, tags)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (user_id, '2026-01-20', '19:45', 2, 'Daily two', 'Body two', 'AI two', 'tag3'),
        )

        conn.execute(
            """
            INSERT INTO dreamdiary_entries
            (user_id, entry_date, entry_time, entry_number, title, cast, location, period,
             emotion, plot, symbols_and_imagery, insight, action, other, dream_places, tags)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                user_id,
                '2026-01-11',
                '08:15',
                1,
                'Dream one',
                'Alex',
                'Forest',
                'Present day',
                'Joy',
                'Flying over trees',
                'Birds and wind',
                'Need freedom',
                'Kept flying',
                'None',
                'Forest',
                'dream,flight',
            ),
        )
        conn.execute(
            """
            INSERT INTO dreamdiary_entries
            (user_id, entry_date, entry_time, entry_number, title, cast, location, period,
             emotion, plot, symbols_and_imagery, insight, action, other, dream_places, tags)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                user_id,
                '2026-01-21',
                '07:50',
                2,
                'Dream two',
                'Sam',
                'Beach',
                'Future',
                'Calm',
                'Walking on water',
                'Moonlight',
                'Trust myself',
                'Kept walking',
                'None',
                'Beach',
                'dream,water',
            ),
        )
        conn.commit()
        conn.close()

    def test_export_unauthenticated(self, client):
        resp = client.get('/api/import/export')
        assert resp.status_code == 401

    def test_export_returns_zip_package_with_expected_sheets_and_headers(self, client):
        token = _register_and_login(client)

        db_path = os.environ['DB_PATH']
        self._seed_export_rows(db_path)

        resp = client.get(
            '/api/import/export',
            headers={'Authorization': f'Bearer {token}'},
        )

        assert resp.status_code == 200
        assert 'zip' in resp.content_type or 'octet-stream' in resp.content_type

        wb, manifest, package_members = _load_export_package(resp.data)
        assert 'Daily' in wb.sheetnames
        assert 'Dreams' in wb.sheetnames
        assert manifest['package_type'] == 'openmynd_export'
        assert manifest['version'] == 1
        assert manifest['portability']['contract_version'] == 1
        assert manifest['portability']['workbook_fields']['daily'] == list(DAILY_IMPORT_HEADERS)
        assert manifest['portability']['workbook_fields']['dream'] == list(DREAM_IMPORT_HEADERS)
        assert 'attachment-derived text and transcripts' in manifest['portability']['omitted_data']
        assert 'entries.xlsx' in package_members
        assert 'manifest.json' in package_members

        daily_ws = wb['Daily']
        dream_ws = wb['Dreams']

        daily_headers = [cell.value for cell in daily_ws[1]]
        dream_headers = [cell.value for cell in dream_ws[1]]
        assert daily_headers == list(DAILY_IMPORT_HEADERS)
        assert dream_headers == list(DREAM_IMPORT_HEADERS)

        assert daily_ws.max_row == 3
        assert dream_ws.max_row == 3

        assert daily_ws.cell(2, 1).value == '2026-01-10'
        assert daily_ws.cell(2, 2).value == '09:30'
        assert daily_ws.cell(2, 3).value == 'Daily one'
        assert daily_ws.cell(2, 4).value == 'Body text'
        assert daily_ws.cell(2, 5).value == 'AI text'

        assert dream_ws.cell(2, 1).value == '2026-01-11'
        assert dream_ws.cell(2, 2).value == '08:15'
        assert dream_ws.cell(2, 3).value == 'Dream one'
        assert dream_ws.cell(2, 4).value == 'Flying over trees'
        assert dream_ws.cell(2, 5).value == 'Alex'
        assert dream_ws.cell(2, 6).value == 'Forest'
        assert dream_ws.cell(2, 13).value == 'dream,flight'

    def test_export_package_round_trips_supported_entry_fields_to_another_user(self, client):
        source_token = _register_and_login(client)
        self._seed_export_rows(os.environ['DB_PATH'])

        export_response = client.get(
            '/api/import/export',
            headers={'Authorization': f'Bearer {source_token}'},
        )
        assert export_response.status_code == 200

        client.post(
            '/api/register',
            data=json.dumps({'username': 'restore-user', 'password': 'secret123'}),
            content_type='application/json',
        )
        login_response = client.post(
            '/api/login',
            data=json.dumps({'username': 'restore-user', 'password': 'secret123'}),
            content_type='application/json',
        )
        restore_token = json.loads(login_response.data)['token']

        import_response = _upload(
            client,
            restore_token,
            export_response.data,
            filename='round-trip.zip',
            content_type='application/zip',
        )
        assert import_response.status_code == 200
        result = json.loads(import_response.data)
        assert result['summary']['inserted_daily'] == 2
        assert result['summary']['inserted_dreams'] == 2
        assert any('portability notice' in warning.lower() for warning in result['warnings'])

        conn = sqlite3.connect(os.environ['DB_PATH'])
        conn.row_factory = sqlite3.Row
        restored_user_id = conn.execute(
            'SELECT id FROM users WHERE username = ?',
            ('restore-user',),
        ).fetchone()['id']
        daily = conn.execute(
            '''
            SELECT entry_date, entry_time, title, user_message, ai_response
            FROM dailydiary_entries
            WHERE user_id = ? AND title = 'Daily one'
            ''',
            (restored_user_id,),
        ).fetchone()
        dream = conn.execute(
            '''
            SELECT entry_date, entry_time, title, plot, "cast", location, emotion, tags
            FROM dreamdiary_entries
            WHERE user_id = ? AND title = 'Dream one'
            ''',
            (restored_user_id,),
        ).fetchone()
        conn.close()

        assert dict(daily) == {
            'entry_date': '2026-01-10',
            'entry_time': '09:30',
            'title': 'Daily one',
            'user_message': 'Body text',
            'ai_response': 'AI text',
        }
        assert {
            key: dream[key]
            for key in ('entry_date', 'entry_time', 'title', 'plot', 'cast', 'location', 'emotion')
        } == {
            'entry_date': '2026-01-11',
            'entry_time': '08:15',
            'title': 'Dream one',
            'plot': 'Flying over trees',
            'cast': 'Alex',
            'location': 'Forest',
            'emotion': 'Joy',
        }
        assert dream['tags'].split(',')[:2] == ['dream', 'flight']

    def test_export_records_no_guard_token_for_manual_date_range_export(self, client):
        token = _register_and_login(client)
        db_path = os.environ['DB_PATH']
        self._seed_export_rows(db_path)

        resp = client.get(
            '/api/import/export?from_date=2026-01-10&to_date=2026-01-21&include_daily=true&include_dreams=true',
            headers={'Authorization': f'Bearer {token}'},
        )

        assert resp.status_code == 200
        guard_token = resp.headers.get('X-OpenMynd-Export-Token')
        assert guard_token is None
        assert resp.headers.get('X-AiDiary-Export-Token') is None

        conn = sqlite3.connect(db_path)
        row = conn.execute(
            '''
            SELECT from_date, to_date, include_daily, include_dreams, is_full_range,
                   daily_count, dream_count, guard_token, used_for_bulk_delete
            FROM export_history
            ORDER BY id DESC
            LIMIT 1
            '''
        ).fetchone()
        conn.close()

        assert row is not None
        assert row[0] == '2026-01-10'
        assert row[1] == '2026-01-21'
        assert row[2] == 1
        assert row[3] == 1
        assert row[4] == 0
        assert row[5] == 2
        assert row[6] == 2
        assert row[7] == guard_token
        assert row[8] == 0

    def test_export_does_not_issue_guard_token_for_partial_export(self, client):
        token = _register_and_login(client)
        self._seed_export_rows(os.environ['DB_PATH'])

        resp = client.get(
            '/api/import/export?from_date=2026-01-11&to_date=2026-01-20',
            headers={'Authorization': f'Bearer {token}'},
        )

        assert resp.status_code == 200
        assert resp.headers.get('X-OpenMynd-Export-Token') is None

    def test_export_daily_only(self, client):
        token = _register_and_login(client)
        self._seed_export_rows(os.environ['DB_PATH'])

        resp = client.get(
            '/api/import/export?include_daily=true&include_dreams=false',
            headers={'Authorization': f'Bearer {token}'},
        )

        assert resp.status_code == 200
        wb, _, _ = _load_export_package(resp.data)
        assert wb.sheetnames == ['Daily']

        ws = wb['Daily']
        headers = [cell.value for cell in ws[1]]
        assert headers == list(DAILY_IMPORT_HEADERS)
        assert ws.max_row == 3

    def test_export_dreams_only(self, client):
        token = _register_and_login(client)
        self._seed_export_rows(os.environ['DB_PATH'])

        resp = client.get(
            '/api/import/export?include_daily=false&include_dreams=true',
            headers={'Authorization': f'Bearer {token}'},
        )

        assert resp.status_code == 200
        wb, _, _ = _load_export_package(resp.data)
        assert wb.sheetnames == ['Dreams']

        ws = wb['Dreams']
        headers = [cell.value for cell in ws[1]]
        assert headers == list(DREAM_IMPORT_HEADERS)
        assert ws.max_row == 3

    def test_export_date_range_filter(self, client):
        token = _register_and_login(client)
        self._seed_export_rows(os.environ['DB_PATH'])

        resp = client.get(
            '/api/import/export?from_date=2026-01-11&to_date=2026-01-20',
            headers={'Authorization': f'Bearer {token}'},
        )

        assert resp.status_code == 200
        wb, _, _ = _load_export_package(resp.data)

        daily_ws = wb['Daily']
        dream_ws = wb['Dreams']

        assert daily_ws.max_row == 2
        assert daily_ws.cell(2, 1).value == '2026-01-20'
        assert dream_ws.max_row == 2
        assert dream_ws.cell(2, 1).value == '2026-01-11'

    def test_export_package_includes_bundled_images_and_manifest_asset_ref(self, client):
        token = _register_and_login(client)
        db_path = os.environ['DB_PATH']
        self._seed_export_rows(db_path)

        media_root = os.environ['MEDIA_ROOT']
        storage_key = 'entries/daily/1/export-test.png'
        absolute_path = os.path.join(media_root, *storage_key.split('/'))
        os.makedirs(os.path.dirname(absolute_path), exist_ok=True)
        with open(absolute_path, 'wb') as image_file:
            image_file.write(_tiny_png_bytes())

        conn = sqlite3.connect(db_path)
        conn.execute(
            """
            UPDATE dailydiary_entries
            SET image_storage_key = ?, image_source = ?, image_position_x = ?, image_position_y = ?, image_prompt = ?
            WHERE title = 'Daily one'
            """,
            (storage_key, 'ai', '44', '62', 'A quiet city street'),
        )
        conn.commit()
        conn.close()

        resp = client.get(
            '/api/import/export',
            headers={'Authorization': f'Bearer {token}'},
        )

        assert resp.status_code == 200
        wb, manifest, package_members = _load_export_package(resp.data)
        daily_ws = wb['Daily']
        asset_ref = daily_ws.cell(2, 6).value

        assert asset_ref
        assert asset_ref in manifest['assets']
        assert manifest['assets'][asset_ref]['image_source'] == 'ai'
        assert manifest['assets'][asset_ref]['image_position_x'] == '44'
        assert manifest['assets'][asset_ref]['image_position_y'] == '62'
        assert any(name.startswith(f'media/{asset_ref}/') for name in package_members)

    def test_export_package_includes_entry_attachments(self, client):
        token = _register_and_login(client)
        db_path = os.environ['DB_PATH']
        self._seed_export_rows(db_path)

        media_root = os.environ['MEDIA_ROOT']
        attachment_key = 'assets/daily/1/receipt.pdf'
        attachment_path = os.path.join(media_root, *attachment_key.split('/'))
        os.makedirs(os.path.dirname(attachment_path), exist_ok=True)
        with open(attachment_path, 'wb') as attachment_file:
            attachment_file.write(b'%PDF-1.4 receipt\n')

        conn = sqlite3.connect(db_path)
        conn.execute(
            """
            INSERT INTO entry_assets
            (user_id, entry_type, entry_id, asset_role, storage_key, original_filename,
             mime_type, file_size_bytes, sort_order, created_at)
            VALUES (?, 'daily', ?, 'attachment', ?, 'receipt.pdf', 'application/pdf', ?, 0, ?)
            """,
            (1, 1, attachment_key, 17, '2026-06-12T10:00:00Z'),
        )
        conn.commit()
        conn.close()

        resp = client.get(
            '/api/import/export',
            headers={'Authorization': f'Bearer {token}'},
        )

        assert resp.status_code == 200
        wb, manifest, package_members = _load_export_package(resp.data)
        daily_ws = wb['Daily']
        asset_ref = daily_ws.cell(2, 6).value

        assert asset_ref
        assert asset_ref in manifest['assets']
        attachments = manifest['assets'][asset_ref]['attachments']
        assert len(attachments) == 1
        assert attachments[0]['original_filename'] == 'receipt.pdf'
        assert attachments[0]['mime_type'] == 'application/pdf'
        assert attachments[0]['package_filename'] == '01_receipt.pdf'
        assert f'media/{asset_ref}/01_receipt.pdf' in package_members

    def test_export_invalid_date_format(self, client):
        token = _register_and_login(client)

        resp = client.get(
            '/api/import/export?from_date=2026/01/10',
            headers={'Authorization': f'Bearer {token}'},
        )

        assert resp.status_code == 400
        data = json.loads(resp.data)
        assert data['status'] == 'error'
        assert 'from_date' in data['errors'][0]

    def test_export_from_date_after_to_date(self, client):
        token = _register_and_login(client)

        resp = client.get(
            '/api/import/export?from_date=2026-01-20&to_date=2026-01-10',
            headers={'Authorization': f'Bearer {token}'},
        )

        assert resp.status_code == 400
        data = json.loads(resp.data)
        assert data['status'] == 'error'
        assert 'cannot be after' in data['errors'][0]

    def test_export_no_types_selected(self, client):
        token = _register_and_login(client)

        resp = client.get(
            '/api/import/export?include_daily=false&include_dreams=false',
            headers={'Authorization': f'Bearer {token}'},
        )

        assert resp.status_code == 400
        data = json.loads(resp.data)
        assert data['status'] == 'error'
        assert 'At least one export type' in data['errors'][0]

    def test_export_all_includes_important_days_and_thought_records(self, client):
        token = _register_and_login(client)
        self._seed_export_rows(os.environ['DB_PATH'])
        conn = sqlite3.connect(os.environ['DB_PATH'])
        user_id = conn.execute("SELECT id FROM users WHERE username = 'importer'").fetchone()[0]
        conn.execute(
            """
            INSERT INTO important_days (
                user_id, label, starts_on, month, day, original_year, category,
                recurrence, icon_name, accent_color, note
            ) VALUES (?, 'Mum letter', '2026-07-20', 7, 20, 2026, 'milestone',
                      'yearly', 'event', 'blue', 'CBT task')
            """,
            (user_id,),
        )
        conn.execute(
            """
            INSERT INTO cbt_worksheets
            (user_id, worksheet_type, title, status, current_step, record_date)
            VALUES (?, 'thought_record', 'Balanced view', 'completed', 7, '2026-07-21')
            """,
            (user_id,),
        )
        worksheet_id = conn.execute('SELECT MAX(id) FROM cbt_worksheets').fetchone()[0]
        conn.execute(
            """
            INSERT INTO cbt_thought_record_data (
                worksheet_id, situation, feelings_before_json, unhelpful_thoughts,
                evidence_for, evidence_against, balanced_thought,
                feelings_after_json, next_step, ai_response
            ) VALUES (?, 'Situation text', '[]', 'Unhelpful', 'For', 'Against',
                      'Balanced', '[]', 'Next', 'AI note')
            """,
            (worksheet_id,),
        )
        conn.commit()
        conn.close()

        resp = client.get(
            '/api/import/export?export_all=true',
            headers={'Authorization': f'Bearer {token}'},
        )

        assert resp.status_code == 200
        assert resp.headers.get('X-OpenMynd-Export-Token')
        workbook, manifest, _ = _load_export_package(resp.data)
        assert workbook['Important Days'].cell(1, 1).value == IMPORTANT_DAY_IMPORT_HEADERS[0]
        assert workbook['Important Days'].cell(2, 2).value == 'Mum letter'
        assert workbook['Thought Records'].cell(1, 1).value == THOUGHT_RECORD_IMPORT_HEADERS[0]
        assert workbook['Thought Records'].cell(2, 2).value == 'Balanced view'
        assert 'important_days' in manifest['portability']['workbook_fields']


def test_zip_package_import_restores_important_days_and_thought_records(client):
    token = _register_and_login(client)
    asset_ref = 'important_day_20260720_1'
    package_bytes = _make_all_data_zip_package(
        important_day_rows=[[
            '2026-07-20',
            'Mum letter',
            'milestone',
            'yearly',
            'event',
            'blue',
            'CBT task',
            asset_ref,
        ]],
        thought_record_rows=[[
            '2026-07-21',
            'Balanced view',
            'completed',
            7,
            '',
            '',
            'Situation text',
            '[]',
            'Unhelpful thought',
            'Evidence for',
            'Evidence against',
            'Balanced thought',
            '[]',
            'Next step',
            'AI response',
            '',
            0,
        ]],
        manifest_assets={
            asset_ref: {
                'entry_type': 'important_day',
                'image_filename': 'important.jpg',
            },
        },
        media_files={
            f'media/{asset_ref}/important.jpg': _tiny_jpeg_bytes(),
        },
    )

    upload_resp = _upload(client, token, package_bytes, filename='openmynd.zip', content_type='application/zip')
    assert upload_resp.status_code == 200
    upload_data = json.loads(upload_resp.data)
    assert upload_data['status'] == 'review_required'
    assert upload_data['summary']['ready_important_days'] == 1
    assert upload_data['summary']['ready_thought_records'] == 1
    selected_row_ids = [row['row_id'] for row in upload_data['review_entries']]

    commit_resp = client.post(
        '/api/import/commit',
        headers={'Authorization': f'Bearer {token}'},
        json={
            'import_session_id': upload_data['import_session_id'],
            'accepted_duplicate_row_ids': [],
            'selected_row_ids': selected_row_ids,
            'entry_type_overrides': {},
        },
    )
    assert commit_resp.status_code == 200
    commit_data = json.loads(commit_resp.data)
    assert commit_data['summary']['inserted_important_days'] == 1
    assert commit_data['summary']['inserted_thought_records'] == 1

    conn = sqlite3.connect(os.environ['DB_PATH'])
    conn.row_factory = sqlite3.Row
    important_day = conn.execute(
        "SELECT label, image_storage_key FROM important_days WHERE label = 'Mum letter'"
    ).fetchone()
    thought_record = conn.execute(
        """
        SELECT w.title, d.balanced_thought, d.ai_response
        FROM cbt_worksheets w
        JOIN cbt_thought_record_data d ON d.worksheet_id = w.id
        WHERE w.title = 'Balanced view'
        """
    ).fetchone()
    conn.close()

    assert important_day['image_storage_key']
    assert thought_record['balanced_thought'] == 'Balanced thought'
    assert thought_record['ai_response'] == 'AI response'
